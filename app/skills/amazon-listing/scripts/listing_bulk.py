#!/usr/bin/env python3
"""Amazon listing flat-file (category template) inspect + fill + feedback.

Drives the "Add Products via Upload" round trip: download a category
template (a macro-enabled `.xlsm` flat file, TemplateType=fptcustom),
fill parent/child rows from a JSON spec, upload it, then parse the
processing report Amazon returns.

Why a script (vs. clicking the listing wizard)
----------------------------------------------
One template upload creates a whole variation family (a Parent plus N
colour/size Children) in a single round trip. The per-field web wizard
is one SKU at a time through shadow-DOM widgets. The flat file is the
supported batch interface -- the same philosophy as `amazon-ads`'
`ads_bulk.py`, but listing templates are **category-specific** (the
column set differs per product type), so this tool cannot use a fixed
positional schema. Instead it keys every field by its **field API name**
(the row that contains `item_sku`), which is identical in every console
language -- only the human label row above it is localised.

Template geometry (verified on a real fptcustom template)
---------------------------------------------------------
  Excel row 1  TemplateType/Version/Signature + group names
  Excel row 2  Local Label Names  (LOCALISED -- never keyed on)
  Excel row 3  field API names    (item_sku, update_delete, ...) <-- keyed
  Excel row 4+ data rows

The operation column is `update_delete`:
  blank  -> Create   (the default when no ASIN is supplied)
  Update / partialupdate / delete
The parent-child cluster is `parent_child` (Parent/Child),
`relationship_type` (Variation), `variation_theme` (e.g. Color),
`parent_sku`. Update-by-ASIN sets `external_product_id` = the ASIN and
`external_product_id_type` = asin.

The template's own sheets are the source of truth for validation:
  'Data Definitions'  -> which fields are Required
  'Valid Values' / 'Dropdown Lists' -> the accepted enum tokens

Fill preserves the workbook verbatim (macros, data validation, the
signature row) and only writes data rows -- Amazon's validator keys on
that untouched header block, so we never rebuild it.

Usage
-----
  listing_bulk.py inspect  TEMPLATE.xlsm [--field NAME]
  listing_bulk.py fill     TEMPLATE.xlsm --spec SPEC.json --out OUT.xlsm
  listing_bulk.py parse-feedback  REPORT[.xlsm|.txt|.csv]

Requires: openpyxl.
"""

import argparse
import csv
import json
import shutil
import sys

try:
    import openpyxl
except ImportError:
    sys.exit('error: openpyxl is required (pip install openpyxl)')


# The Template sheet holds the flat file. Metadata lives in siblings.
TEMPLATE_SHEET = 'Template'
DEFN_SHEET = 'Data Definitions'
VALID_SHEET = 'Valid Values'
DROPDOWN_SHEET = 'Dropdown Lists'

# The operation column and the tokens Amazon accepts in it. A blank
# cell means Create -- that is the default and cannot be a literal
# token, so callers pass operation='create' and we clear the cell.
OP_COLUMN = 'update_delete'
OP_TOKENS = {
    'create': '',
    'update': 'Update',
    'partialupdate': 'partialupdate',
    'delete': 'delete',
}

# Fields the caller addresses through friendly per-row keys, mapped to
# their flat-file field API name. Everything else goes in `fields`.
IDENTITY_FIELD = 'item_sku'
PRODUCT_TYPE_FIELD = 'feed_product_type'
BRAND_FIELD = 'brand_name'

# A Parent (relationship) row carries only the shared catalogue data;
# the offer, stock, images and product-id live on the Child rows. So a
# Parent legitimately omits these even when the template marks them
# Required -- don't warn about them for a Parent.
CHILD_LEVEL_PREFIXES = (
    'purchasable_offer',
    'fulfillment_availability',
    'main_image_url',
    'other_image_url',
    'swatch_image_url',
    'external_product_id',
    'color_name',
    'size_name',
    'apparel_size',
)

# Battery / hazmat / dangerous-goods fields are marked Required in the
# Data Definitions but are only *conditionally* required -- Amazon
# accepts them blank when the item has no batteries. Suppress the noise
# when the row declares no batteries.
BATTERY_TOKENS = (
    'battery',
    'batteries',
    'lithium',
    'hazmat',
    'ghs_classification',
    'safety_data_sheet',
    'flash_point',
    'supplier_declared_dg',
    'number_of_lithium',
)


def _is_falsey(v):
    return v is None or str(v).strip().lower() in ('', 'no', 'false', '0')


def _keep_vba(path):
    """Real Amazon templates are macro-enabled .xlsm (keep the macros and
    the signature row); tolerate a plain .xlsx too (tests, blank forms)."""
    return path.lower().endswith('.xlsm')


def _find_header_row(ws, max_scan=8):
    """Return the 1-based row index whose cells carry field API names.

    Identified structurally by the presence of `item_sku`, so it works
    regardless of console language (the label row above it is localised;
    this row is not).
    """
    for r in range(1, max_scan + 1):
        values = [c.value for c in ws[r]]
        if IDENTITY_FIELD in values:
            return r
    raise ValueError(
        f"could not find the field-name row (no '{IDENTITY_FIELD}' in the "
        f'first {max_scan} rows) -- is this a listing flat-file template?'
    )


def _field_columns(ws, header_row):
    """Map field API name -> list of 1-based column indices.

    Repeated fields (bullet_point1..N are distinct, but some templates
    reuse a bare name across marketplace-scoped offer blocks) map to
    multiple columns; the first is used for scalar writes.
    """
    cols = {}
    for cell in ws[header_row]:
        name = cell.value
        if name:
            cols.setdefault(str(name), []).append(cell.column)
    return cols


def _load_required_fields(wb):
    """Field API names marked Required in the Data Definitions sheet.

    Data Definitions layout: header at Excel row 2, then per-field rows
    with columns Group Name | Field Name | Local Label | Definition |
    Accepted Values | Example | Required?.
    """
    if DEFN_SHEET not in wb.sheetnames:
        return set()
    ws = wb[DEFN_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    required = set()
    for row in rows[2:]:
        if not row or len(row) < 7:
            continue
        field_name, req = row[1], row[6]
        if field_name and req and str(req).strip().lower() == 'required':
            required.add(str(field_name).strip())
    return required


def _load_valid_values(wb):
    """Map field API name -> set of accepted enum tokens (lowercased).

    Read from 'Dropdown Lists', whose row 3 holds field API names as
    column headers and the rows below hold the tokens for each. Returns
    {} silently if the sheet is absent or shaped unexpectedly -- enum
    checks then degrade to warnings only.
    """
    if DROPDOWN_SHEET not in wb.sheetnames:
        return {}
    ws = wb[DROPDOWN_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        return {}
    # The field-name header inside Dropdown Lists is the row that
    # contains a known enum field; probe the first few rows for it.
    hdr_idx = None
    for i in range(min(5, len(rows))):
        if rows[i] and OP_COLUMN in [str(c) for c in rows[i] if c]:
            hdr_idx = i
            break
    if hdr_idx is None:
        return {}
    hdr = rows[hdr_idx]
    out = {}
    for ci, name in enumerate(hdr):
        if not name:
            continue
        vals = set()
        for ri in range(hdr_idx + 1, len(rows)):
            v = rows[ri][ci] if ci < len(rows[ri]) else None
            if v not in (None, ''):
                vals.add(str(v).strip().lower())
        if vals:
            out[str(name).strip()] = vals
    return out


def _valid_value_case(wb):
    """Map field API name -> {lowercased token: template's exact-case token}.

    Amazon rejects a case-mismatched enum on some fields (verified live:
    `apparel_size_system` accepts `UAE/KSA` but rejects `uae/ksa`), so a
    spec value must be written in the template's own case. This mirrors
    `_load_valid_values` but keeps the original casing as the value.
    """
    if DROPDOWN_SHEET not in wb.sheetnames:
        return {}
    ws = wb[DROPDOWN_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        return {}
    hdr_idx = None
    for i in range(min(5, len(rows))):
        if rows[i] and OP_COLUMN in [str(c) for c in rows[i] if c]:
            hdr_idx = i
            break
    if hdr_idx is None:
        return {}
    out = {}
    for ci, name in enumerate(rows[hdr_idx]):
        if not name:
            continue
        m = {}
        for ri in range(hdr_idx + 1, len(rows)):
            v = rows[ri][ci] if ci < len(rows[ri]) else None
            if v not in (None, ''):
                m[str(v).strip().lower()] = str(v).strip()
        if m:
            out[str(name).strip()] = m
    return out


def _export_tsv(ws, path):
    """Write the Template sheet as a tab-delimited .txt (the upload file).

    Amazon's flat-file processor rejects an openpyxl-saved `.xlsm` with a
    90502 FATAL ("worksheet template type not supported for Excel
    upload") because the re-save alters the macro-workbook structure it
    validates. Its own remedy is to upload a tab-delimited text file, so
    this is the artefact you actually upload -- keyed by the row-3 field
    names, uniform column count so nothing shifts.
    """
    max_c = ws.max_column
    with open(path, 'w', newline='', encoding='utf-8') as fh:
        w = csv.writer(fh, delimiter='\t', lineterminator='\n')
        for row in ws.iter_rows():
            w.writerow(
                ['' if c.value is None else c.value for c in row][:max_c]
            )


def cmd_inspect(args):
    wb = openpyxl.load_workbook(
        args.file, read_only=False, keep_vba=_keep_vba(args.file)
    )
    ws = wb[TEMPLATE_SHEET]
    header_row = _find_header_row(ws)
    cols = _field_columns(ws, header_row)
    required = _load_required_fields(wb)
    valid = _load_valid_values(wb)

    def dig(name):
        v = valid.get(name)
        return sorted(v) if v else None

    if args.field:
        name = args.field
        print(f'field: {name}')
        print(f'  columns : {cols.get(name)}')
        print(f'  required: {name in required}')
        print(f'  values  : {dig(name)}')
        return

    tt = ws.cell(row=1, column=1).value
    print(f'template : {tt}')
    print(f'sheets   : {wb.sheetnames}')
    print(f'header at Excel row {header_row}; data starts row {header_row + 1}')
    print(f'total fields: {len(cols)}   required fields: {len(required)}')
    print(
        '\noperation column ({}): create=blank / {}'.format(
            OP_COLUMN, ' / '.join(t for t in OP_TOKENS.values() if t)
        )
    )
    print('\nvariation cluster:')
    for f in (
        'parent_child',
        'relationship_type',
        'variation_theme',
        'parent_sku',
        'external_product_id',
        'external_product_id_type',
    ):
        if f in cols:
            print(f'  {f:26} col={cols[f][0]:<4} values={dig(f)}')
    print('\nrequired fields:')
    for f in sorted(required):
        print(f'  {f:30} values={dig(f)}')


def _resolve_operation(op):
    key = (op or 'create').strip().lower()
    if key not in OP_TOKENS:
        raise SystemExit(
            f"error: operation '{op}' is not one of {', '.join(OP_TOKENS)}"
        )
    return OP_TOKENS[key], key


def _row_fields(spec_row, top):
    """Flatten one spec row into {field_api_name: value}.

    Friendly per-row keys (sku, asin, parent_sku, parentage,
    variation_theme) fold into their flat-file field names; `fields`
    carries everything else verbatim by API name. Top-level `product_type`
    and `brand` supply defaults when a row omits them.
    """
    out = dict(spec_row.get('fields') or {})
    out.setdefault(PRODUCT_TYPE_FIELD, top.get('product_type'))
    if top.get('brand'):
        out.setdefault(BRAND_FIELD, top['brand'])
    if spec_row.get('sku'):
        out[IDENTITY_FIELD] = spec_row['sku']
    if spec_row.get('parent_sku'):
        out['parent_sku'] = spec_row['parent_sku']
    if spec_row.get('parentage'):
        out['parent_child'] = spec_row['parentage']
    if spec_row.get('variation_theme'):
        out['variation_theme'] = spec_row['variation_theme']
    if spec_row.get('asin'):
        out['external_product_id'] = spec_row['asin']
        out.setdefault('external_product_id_type', 'asin')
    # Drop keys with no value so we never blank an intended default.
    return {k: v for k, v in out.items() if v not in (None, '')}


def cmd_fill(args):
    with open(args.spec, encoding='utf-8') as fh:
        spec = json.load(fh)
    rows = spec.get('rows') or []
    if not rows:
        raise SystemExit('error: spec has no rows')

    shutil.copyfile(args.file, args.out)
    wb = openpyxl.load_workbook(
        args.out, read_only=False, keep_vba=_keep_vba(args.out)
    )
    ws = wb[TEMPLATE_SHEET]
    header_row = _find_header_row(ws)
    cols = _field_columns(ws, header_row)
    required = _load_required_fields(wb)
    valid = _load_valid_values(wb)
    case = _valid_value_case(wb)

    unknown_fields = set()
    warnings = []
    write_at = header_row + 1
    # Clear any pre-existing data rows before writing the spec, so stale
    # rows from an Example row or a previously-filled template can't be
    # uploaded as unintended SKUs. A freshly-generated blank template has
    # none, but a reused workbook might.
    if ws.max_row >= write_at:
        ws.delete_rows(write_at, ws.max_row - write_at + 1)
    for i, spec_row in enumerate(rows):
        op_token, op_key = _resolve_operation(spec_row.get('operation'))
        fields = _row_fields(spec_row, spec)
        sku = fields.get(IDENTITY_FIELD)
        if not sku:
            raise SystemExit(f'error: row {i} has no sku')

        # Enum validation: reject an invalid operation-family token hard;
        # warn (never fail) on other enums so an unseen-but-valid token
        # from a new category still uploads.
        for fname, fval in fields.items():
            allowed = valid.get(fname)
            if allowed and str(fval).strip().lower() not in allowed:
                warnings.append(
                    f'row {i} sku={sku}: {fname}={fval!r} not in template '
                    f'valid values {sorted(allowed)}'
                )

        # Required-field guard applies to Create rows only. Update/
        # partialupdate touch a subset; delete needs just sku+operation.
        if op_key == 'create':
            is_parent = str(fields.get('parent_child', '')).lower() == 'parent'
            no_battery = _is_falsey(
                fields.get('are_batteries_included')
            ) and _is_falsey(fields.get('batteries_required'))
            missing = []
            for f in required:
                if f not in cols or f in fields or f == OP_COLUMN:
                    continue
                if is_parent and f.startswith(CHILD_LEVEL_PREFIXES):
                    continue
                if no_battery and any(t in f for t in BATTERY_TOKENS):
                    continue
                missing.append(f)
            if missing:
                warnings.append(
                    f'row {i} sku={sku}: missing required field(s) '
                    f'{sorted(missing)}'
                )

        target = ws[write_at + i]
        # Operation column: write the token (blank cell = Create).
        if OP_COLUMN in cols:
            target[cols[OP_COLUMN][0] - 1].value = op_token or None
        for fname, fval in fields.items():
            if fname not in cols:
                unknown_fields.add(fname)
                continue
            # Canonicalise to the template's exact-case enum token when
            # this field has a known valid set -- some fields are
            # case-strict on Amazon's side (e.g. UAE/KSA vs uae/ksa).
            canon = case.get(fname, {}).get(str(fval).strip().lower())
            target[cols[fname][0] - 1].value = canon if canon else fval

    wb.save(args.out)
    # The upload artefact is a tab-delimited .txt, NOT this .xlsm (an
    # openpyxl-saved .xlsm triggers Amazon's 90502 FATAL). Write it next
    # to the .xlsm so the caller uploads the .txt.
    txt_path = args.out.rsplit('.', 1)[0] + '.txt'
    _export_tsv(ws, txt_path)

    for w in warnings:
        print(f'warning: {w}', file=sys.stderr)
    if unknown_fields:
        print(
            f'warning: fields not in this template (skipped): '
            f'{sorted(unknown_fields)}',
            file=sys.stderr,
        )
    print(f'wrote {len(rows)} data row(s) -> {args.out}')
    print(
        f'UPLOAD THIS (tab-delimited) -> {txt_path}  '
        '(uploading the Excel file itself triggers a 90502 FATAL)'
    )


def _iter_report_rows(path):
    """Yield rows from a processing report (.xlsm/.xlsx or .txt/.csv)."""
    lower = path.lower()
    if lower.endswith(('.xlsm', '.xlsx')):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for name in wb.sheetnames:
            for row in wb[name].iter_rows(values_only=True):
                yield [('' if c is None else str(c)) for c in row]
        return
    delim = '\t'
    with open(path, encoding='utf-8', errors='replace') as fh:
        sample = fh.readline()
        if ',' in sample and '\t' not in sample:
            delim = ','
        fh.seek(0)
        for line in fh:
            yield line.rstrip('\n').split(delim)


def _template_cell_errors(path):
    """Per-field errors from a processing report's Template tab.

    Amazon writes the precise, field-level fix as CELL COMMENTS (批注) on
    the report's Template tab: the summary table gives code + message,
    but the comment pins the exact column (field) to change. read_only
    mode does not load comments, so open the workbook normally. Yields
    (sku, field_api_name, comment_text) -- the self-correction targets.
    """
    if not path.lower().endswith(('.xlsm', '.xlsx')):
        return
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        if TEMPLATE_SHEET not in wb.sheetnames:
            return
        ws = wb[TEMPLATE_SHEET]
        try:
            header_row = _find_header_row(ws)
        except ValueError:
            return
        names = {
            c.column: str(c.value).strip() for c in ws[header_row] if c.value
        }
        sku_col = next(
            (col for col, n in names.items() if n == IDENTITY_FIELD), None
        )
        for row in ws.iter_rows(min_row=header_row + 1):
            sku = ''
            if sku_col is not None:
                cell = next((c for c in row if c.column == sku_col), None)
                sku = cell.value if cell and cell.value else ''
            for c in row:
                if not (c.comment and c.comment.text):
                    continue
                field = names.get(c.column, f'col{c.column}')
                # A cell comment can stack several messages, separated by
                # Excel's literal carriage-return marker `_x000d_`. Split
                # into individual ERROR/WARNING lines so each is one
                # actionable item.
                raw = str(c.comment.text).replace('_x000d_', '\n')
                for line in raw.split('\n'):
                    line = ' '.join(line.split())
                    if line:
                        yield (str(sku or '?'), field, line)
    finally:
        wb.close()


def cmd_parse_feedback(args):
    """Summarise Amazon's processing report: per-SKU errors/warnings.

    Report layouts vary (a summary header block then a table). We locate
    the table header by finding a row that names an error/SKU column,
    then emit each data row's SKU + type + message. Falls back to
    printing every row that mentions 'error' if no header is found.
    """
    rows = list(_iter_report_rows(args.file))
    if not rows:
        raise SystemExit('error: empty report')

    def idx_of(header, *needles):
        for j, cell in enumerate(header):
            c = cell.strip().lower()
            if any(n in c for n in needles):
                return j
        return None

    hdr_i = None
    for i, row in enumerate(rows):
        joined = ' '.join(row).lower()
        if ('error' in joined or 'sku' in joined) and (
            'sku' in joined or 'record' in joined
        ):
            hdr_i = i
            break

    n_err = n_warn = 0
    if hdr_i is not None:
        header = rows[hdr_i]
        c_sku = idx_of(header, 'sku')
        c_type = idx_of(header, 'error type', 'type', 'severity')
        c_msg = idx_of(header, 'message', 'error message', 'description')
        c_code = idx_of(header, 'code')
        print(f'report table header at row {hdr_i}: {header}')
        for row in rows[hdr_i + 1 :]:
            if not any(cell.strip() for cell in row):
                continue
            sku = row[c_sku] if c_sku is not None and c_sku < len(row) else ''
            typ = (
                row[c_type] if c_type is not None and c_type < len(row) else ''
            )
            msg = row[c_msg] if c_msg is not None and c_msg < len(row) else ''
            code = (
                row[c_code] if c_code is not None and c_code < len(row) else ''
            )
            tl = typ.strip().lower()
            if tl == 'error' or 'error' in tl:
                n_err += 1
            elif 'warn' in tl:
                n_warn += 1
            print(f'  [{typ or "?"}] sku={sku} code={code}: {msg}')
    else:
        print('(no structured table found; rows mentioning "error")')
        for row in rows:
            if 'error' in ' '.join(row).lower():
                print('  ' + ' | '.join(row))
                n_err += 1

    # The precise, field-level fixes live in the Template-tab cell
    # comments (批注). Print them as (sku, field, message) -- for each,
    # set that field to a valid value (see `inspect`) and re-upload. The
    # error set is category-specific and unbounded, so extract it from
    # the report rather than hardcode fixes.
    cell_errs = list(_template_cell_errors(args.file))
    if cell_errs:
        print(
            '\nper-field cell comments (批注) -- fix the exact field each names:'
        )
        for sku, field, msg in cell_errs:
            print(f'  sku={sku} field={field}: {msg}')
        # The 批注 are the authoritative per-field verdict; the summary
        # table above is often mislocated, so count from these instead.
        c_err = sum(
            1 for _, _, m in cell_errs if m.lstrip().upper().startswith('ERROR')
        )
        c_warn = sum(
            1 for _, _, m in cell_errs if m.lstrip().upper().startswith('WARN')
        )
        if c_err or c_warn:
            n_err, n_warn = c_err, c_warn

    print(f'\nsummary: {n_err} error(s), {n_warn} warning(s)')
    if n_err:
        sys.exit(1)


def main():
    ap = argparse.ArgumentParser(description=__doc__.split('\n')[0])
    sub = ap.add_subparsers(dest='cmd', required=True)

    p = sub.add_parser('inspect', help='dump fields / required / enums')
    p.add_argument('file')
    p.add_argument('--field', help='detail for one field API name')
    p.set_defaults(func=cmd_inspect)

    p = sub.add_parser('fill', help='write parent/child rows from a spec')
    p.add_argument('file', help='the category template (.xlsm)')
    p.add_argument('--spec', required=True, help='JSON spec of rows')
    p.add_argument('--out', required=True, help='output .xlsm path')
    p.set_defaults(func=cmd_fill)

    p = sub.add_parser('parse-feedback', help='summarise a processing report')
    p.add_argument('file')
    p.set_defaults(func=cmd_parse_feedback)

    args = ap.parse_args()
    args.func(args)


if __name__ == '__main__':
    main()
