#!/usr/bin/env python3
"""Amazon listing flat-file (category template) inspect + fill + feedback.

Drives the "Add Products via Upload" round trip: download a category
template (a macro-enabled `.xlsm` flat file), fill parent/child rows from
a JSON spec, upload it, then parse the processing report Amazon returns.

Why a script (vs. clicking the listing wizard)
----------------------------------------------
One template upload creates a whole variation family (a Parent plus N
colour/size Children) in a single round trip. The per-field web wizard
is one SKU at a time through shadow-DOM widgets. The flat file is the
supported batch interface -- the same philosophy as `amazon-ads`'
`ads_bulk.py`, but listing templates are **category-specific** (the
column set differs per product type), so this tool cannot use a fixed
positional schema. Instead it keys every field by its **field API name**
(the row that contains the SKU column), which is identical in every
console language -- only the human label row above it is localised.

Two template dialects
---------------------
Amazon ships two flat-file templates and the tool auto-detects which:

  legacy  (TemplateType=fptcustom) -- the classic flat file. Field API
          names are bare: `item_sku`, `update_delete`, `parent_child`,
          `relationship_type`, `variation_theme`, `parent_sku`,
          `external_product_id[_type]`.
  unified (NGS "Beta Product Spreadsheet") -- the current Seller Central
          template. Every field name is decorated: the SKU is
          `contribution_sku#1.value`, the operation is `::record_action`,
          parentage is `parentage_level[marketplace_id=<id>]#1.value`, the
          parent link is
          `child_parent_sku_relationship[marketplace_id=<id>]#1.parent_sku`,
          the theme is `variation_theme#1.name`, and the product id lives
          in `amzn1.volt.ca.product_id{_type,_value}`.

Every field the tool addresses by a friendly role (sku, operation,
product_type, brand, parentage, parent_sku, variation_theme, product id,
offer price) is resolved STRUCTURALLY from the header (`_Schema`), so the
same spec drives either dialect. Fields not covered by a role are still
addressed by their exact API name via the spec's `fields` map.

Template geometry
-----------------
  legacy : row 1 TemplateType/signature | row 2 localised labels |
           row 3 field API names <-- keyed | row 4+ data
  unified: rows 1-4 settings/instructions/group/localised labels |
           row 5 field API names <-- keyed | rows 6..dataRow-1 a SKIPPED
           example region | data from `dataRow` (the row-1 settings blob's
           `dataRow=N`, e.g. 8)
The field-name row is found structurally (the row carrying the SKU field);
data is written at `dataRow` (legacy: field-name row + 1), never by a
guessed fixed number.

The operation column (`update_delete` legacy / `::record_action`
unified):
  blank -> Create (the default when no ASIN is supplied)
  update / partialupdate / delete map to each dialect's own tokens.
Update-by-ASIN sets the product-id field to the ASIN and its type to
`asin`.

The template's own sheets are the source of truth for validation:
  'Data Definitions'  -> which fields are Required
  'Valid Values' / 'Dropdown Lists' -> the accepted enum tokens

Fill preserves the workbook verbatim (macros, data validation, the
signature row) and only writes data rows -- Amazon's validator keys on
that untouched header block, so we never rebuild it.

Usage
-----
  listing_bulk.py inspect  TEMPLATE.xlsm [--field NAME]
  listing_bulk.py fill     TEMPLATE.xlsm --spec SPEC.json --out OUT.xlsm
  listing_bulk.py parse-feedback  REPORT[.xlsm|.txt|.csv]

Requires: openpyxl.
"""

import argparse
import csv
import json
import os
import shutil
import sys

try:
    import openpyxl
except ImportError:
    sys.exit('error: openpyxl is required (pip install openpyxl)')

# Global marketplace table + resolution, kept in a sibling data module.
# The script dir is on sys.path when run as a CLI; add it for path-based
# imports (tests) too.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
# Template structure + metadata parsing live in a sibling module so this
# file stays within the line cap. Public names there; alias to the
# `_`-prefixed internal names used here (and re-exported for tests).
from listing_schema import (  # noqa: E402, F401
    DEFN_SHEET,
    DROPDOWN_SHEET,
    OFFER_PRICE_SHORTHANDS as _OFFER_PRICE_SHORTHANDS,
    OP_TOKENS as _OP_TOKENS,
    ROLE_MATCHERS as _ROLE_MATCHERS,
    TEMPLATE_SHEET,
    Schema as _Schema,
    base_attr as _base_attr,
    data_start_row as _data_start_row,
    field_columns as _field_columns,
    find_header_row as _find_header_row,
    is_sku_field as _is_sku_field,
    load_required_fields as _load_required_fields,
    load_valid_values as _load_valid_values,
    our_price_col as _our_price_col,
    route_offer_price as _route_offer_price,
    valid_value_case as _valid_value_case,
)
from marketplace_ids import (  # noqa: E402,F401
    COUNTRY_BY_ID as _COUNTRY_BY_ID,
    MARKETPLACE_IDS,  # re-exported for callers/tests
    fulfillment_index as _fulfillment_index_for_marketplace,
    ids_in_template as _marketplace_ids_in_template,
    resolve as _resolve_marketplace_id,
)


def _mkt_label(mkt_id):
    """'A2VIGQ35RCS4UG (AE)'-style label for messages."""
    cc = _COUNTRY_BY_ID.get(mkt_id)
    return f'{mkt_id} ({cc})' if cc else str(mkt_id)


# Item Highlight (`title_differentiation`) is an OPTIONAL field Amazon only
# accepts when the Item Name is <= 75 chars; a longer title makes Amazon
# reject the highlight with error 100476 ("Provide an Item Name that is 75
# characters or less to use Item Highlights") -- a non-blocking SUCCESS
# (OTHER) error that leaves the SKU in inventory but flagged "Action
# required". The marketing item_name is routinely > 75 chars, so a spec
# that fills Item Highlight (e.g. with the colour) silently poisons every
# child. `fill` drops the optional highlight when the title is too long so
# the SKU lands clean; the value belongs in `color_name`, never here.
_ITEM_HIGHLIGHT_ATTR = 'title_differentiation'
_ITEM_NAME_ATTR = 'item_name'
_ITEM_HIGHLIGHT_MAX_TITLE = 75

# A Parent row carries only shared catalogue data; offer, stock, images
# and product-id live on Child rows -- so a Parent legitimately omits
# these even when the template marks them Required (don't warn). Prefixes
# cover both dialects (`color_name`/legacy id vs `color`/`amzn1.volt.ca`).
CHILD_LEVEL_PREFIXES = (
    'purchasable_offer',
    'fulfillment_availability',
    'main_image_url',
    'main_product_image',
    'other_image_url',
    'other_product_image',
    'swatch_image_url',
    'external_product_id',
    'amzn1.volt.ca.product_id',
    'color_name',
    'color[',
    'size_name',
    'size[',
    'apparel_size',
)

# Battery / hazmat fields are marked Required but are only *conditionally*
# required -- Amazon accepts them blank when the item has no batteries, so
# suppress the noise when the row declares none.
BATTERY_TOKENS = (
    'battery',
    'batteries',
    'lithium',
    'hazmat',
    'ghs_classification',
    'safety_data_sheet',
    'flash_point',
    'supplier_declared_dg',
    'number_of_lithium',
)


def _is_falsey(v):
    return v is None or str(v).strip().lower() in ('', 'no', 'false', '0')


def _keep_vba(path):
    """Real Amazon templates are macro-enabled .xlsm (keep the macros and
    the signature row); tolerate a plain .xlsx too (tests, blank forms)."""
    return path.lower().endswith('.xlsm')


def _export_tsv(ws, path):
    """Write the Template sheet as a tab-delimited .txt (the upload file).

    Amazon's flat-file processor rejects an openpyxl-saved `.xlsm` with a
    90502 FATAL ("worksheet template type not supported for Excel
    upload") because the re-save alters the macro-workbook structure it
    validates. Its own remedy is to upload a tab-delimited text file, so
    this is the artefact you actually upload -- the WHOLE sheet including
    the settings/signature/label header block (Amazon's introspect-feed
    keys on it to detect the file type), uniform column count so nothing
    shifts.
    """
    max_c = ws.max_column
    with open(path, 'w', newline='', encoding='utf-8') as fh:
        w = csv.writer(fh, delimiter='\t', lineterminator='\n')
        for row in ws.iter_rows():
            w.writerow(
                ['' if c.value is None else c.value for c in row][:max_c]
            )


def cmd_inspect(args):
    wb = openpyxl.load_workbook(
        args.file, read_only=False, keep_vba=_keep_vba(args.file)
    )
    ws = wb[TEMPLATE_SHEET]
    header_row = _find_header_row(ws)
    cols = _field_columns(ws, header_row)
    schema = _Schema(cols)
    required = _load_required_fields(wb)
    valid = _load_valid_values(wb)

    def dig(name):
        v = valid.get(name)
        return sorted(v) if v else None

    if args.field:
        name = args.field
        print(f'field: {name}')
        print(f'  columns : {cols.get(name)}')
        print(f'  required: {name in required}')
        print(f'  values  : {dig(name)}')
        return

    tt = str(ws.cell(row=1, column=1).value or '')
    op_field = schema.field('operation')
    data_row = _data_start_row(ws, header_row)
    print(f'template : {tt[:80]}')
    print(f'dialect  : {schema.dialect}')
    mkts = _marketplace_ids_in_template(cols)
    if mkts:
        print(
            'marketplaces: '
            + ', '.join(_mkt_label(m) for m in mkts)
            + '   <- the template is REGION-STAMPED for these; verify '
            'this is the marketplace you intend to list on BEFORE '
            'filling'
        )
    print(f'sheets   : {wb.sheetnames}')
    print(f'header at Excel row {header_row}; data starts row {data_row}')
    print(f'total fields: {len(cols)}   required fields: {len(required)}')
    print(
        '\noperation column ({}): create=blank / {}'.format(
            op_field,
            ' / '.join(t for t in schema.op_tokens().values() if t),
        )
    )
    # The resolved logical roles: the actual columns a spec's friendly keys
    # (and offer-price routing) write to for THIS template's dialect.
    print('\nresolved roles (friendly key -> the column it fills):')
    for role in _ROLE_MATCHERS:
        name = schema.field(role)
        if name:
            print(f'  {role:16} -> {name}   values={dig(name)}')
    print('\nrequired fields:')
    for f in sorted(required):
        print(f'  {f:30} values={dig(f)}')


def _resolve_operation(op, dialect):
    tokens = _OP_TOKENS[dialect]
    key = (op or 'create').strip().lower()
    if key not in tokens:
        raise SystemExit(
            f"error: operation '{op}' is not one of {', '.join(tokens)}"
        )
    return tokens[key], key


def _row_fields(spec_row, top, schema):
    """Flatten one spec row into {field_api_name: value}.

    Friendly per-row keys (sku, asin, parent_sku, parentage,
    variation_theme) fold into their flat-file field names -- resolved
    through `schema` so the SAME spec drives either dialect; `fields`
    carries everything else verbatim by API name. Top-level `product_type`
    and `brand` supply defaults when a row omits them.
    """
    out = dict(spec_row.get('fields') or {})

    def put(role, value, overwrite=True):
        name = schema.field(role)
        if name and value not in (None, ''):
            if overwrite or name not in out:
                out[name] = value

    put('product_type', top.get('product_type'), overwrite=False)
    put('brand', top.get('brand'), overwrite=False)
    put('sku', spec_row.get('sku'))
    put('parent_sku', spec_row.get('parent_sku'))
    put('parentage', spec_row.get('parentage'))
    put('variation_theme', spec_row.get('variation_theme'))
    # A legacy variation row MUST carry relationship_type or a child errors
    # "relationship_type = null" and never creates (explicit wins). The
    # unified template has no such column -- the parent link is carried by
    # parentage_level + child_parent_sku_relationship -- so only add it when
    # the template actually has a `relationship_type` column.
    if (spec_row.get('parentage') or spec_row.get('variation_theme')) and (
        'relationship_type' in schema.cols
    ):
        out.setdefault('relationship_type', 'Variation')
    if spec_row.get('asin'):
        put('product_id', spec_row['asin'])
        put('product_id_type', 'asin', overwrite=False)
    # Offer shorthands (`our_price`/`price`/`quantity`) belong in `fields`,
    # but the skill tells the agent to put "a bare our_price/quantity on
    # each child" -- naturally read as a ROW-LEVEL key. Fold those from the
    # row into fields (a value already in `fields` wins) so the price/stock
    # routes to the target marketplace either way, instead of being
    # silently dropped -> an empty offer column the agent then hand-picks.
    for k in (*_OFFER_PRICE_SHORTHANDS, 'quantity', 'fulfillment_channel_code'):
        if spec_row.get(k) not in (None, '') and k not in out:
            out[k] = spec_row[k]
    # Drop keys with no value so we never blank an intended default.
    return {k: v for k, v in out.items() if v not in (None, '')}


def _drop_unusable_item_highlight(fields, i, sku, warnings):
    """Drop an optional Item Highlight when the Item Name is too long.

    Amazon only accepts `title_differentiation` (Item Highlight) when the
    row's `item_name` is <= 75 chars; otherwise it returns error 100476 and
    parks the SKU in "Action required" (SUCCESS OTHER). The field is
    OPTIONAL, so when the title exceeds the limit we drop the highlight and
    warn rather than upload a value that guarantees a rejection. Mutates
    ``fields``. No-op when no highlight is set or the title fits.
    """
    hi = next(
        (k for k in fields if _base_attr(k) == _ITEM_HIGHLIGHT_ATTR), None
    )
    if hi is None or not str(fields[hi]).strip():
        return
    name = next(
        (v for k, v in fields.items() if _base_attr(k) == _ITEM_NAME_ATTR), ''
    )
    if len(str(name)) > _ITEM_HIGHLIGHT_MAX_TITLE:
        dropped = fields.pop(hi)
        warnings.append(
            f'row {i} sku={sku}: dropped Item Highlight '
            f'({hi}={dropped!r}) -- item_name is '
            f'{len(str(name))} chars (> {_ITEM_HIGHLIGHT_MAX_TITLE}), which '
            f'Amazon rejects with error 100476. Item Highlight is optional; '
            f'put the colour in color_name, not here.'
        )


def cmd_fill(args):
    with open(args.spec, encoding='utf-8') as fh:
        spec = json.load(fh)
    rows = spec.get('rows') or []
    if not rows:
        raise SystemExit('error: spec has no rows')

    shutil.copyfile(args.file, args.out)
    wb = openpyxl.load_workbook(
        args.out, read_only=False, keep_vba=_keep_vba(args.out)
    )
    ws = wb[TEMPLATE_SHEET]
    header_row = _find_header_row(ws)
    cols = _field_columns(ws, header_row)
    schema = _Schema(cols)
    op_field = schema.field('operation')
    required = _load_required_fields(wb)
    valid = _load_valid_values(wb)
    case = _valid_value_case(wb)

    # The offer price is per-marketplace; resolve the target once (CLI
    # --marketplace wins, else spec's top-level "marketplace", else the
    # template itself if it names exactly one marketplace).
    template_ids = _marketplace_ids_in_template(cols)
    requested = getattr(args, 'marketplace', None) or spec.get('marketplace')
    mkt_id = _resolve_marketplace_id(requested, template_ids)
    # Region-stamp guard. A template generated with the wrong store
    # ticked is stamped for the wrong marketplace, and every later step
    # then faithfully "succeeds" on the wrong storefront (observed
    # live). Declaring the target makes that a hard error here; not
    # declaring it on a single-stamped template auto-adopts the stamp,
    # so say LOUDLY which marketplace is about to receive the listing.
    if template_ids and requested and mkt_id not in template_ids:
        raise SystemExit(
            f'error: you are filling for {_mkt_label(mkt_id)} but this '
            'template is region-stamped for '
            + ', '.join(_mkt_label(m) for m in template_ids)
            + ' — it has no offer columns for your target. The template '
            'was generated with the wrong store ticked. Do NOT fill or '
            'upload it; regenerate with the target store ticked '
            '(bh_download_template verifies the tick) and fill that one.'
        )
    if template_ids and not requested and mkt_id:
        print(
            f'warning: no marketplace declared — auto-adopting the '
            f"template's own stamp {_mkt_label(mkt_id)}. The listing "
            'will land on THAT storefront; if that is not the intended '
            'target, STOP and regenerate the template. Declare '
            '--marketplace <CC> (or spec "marketplace") to make a '
            'mismatch fail instead.',
            file=sys.stderr,
        )

    unknown_fields = set()
    warnings = []
    # Data begins at the template's data row (unified: the `dataRow=N` the
    # settings blob names, with rows header_row+1..N-1 an example region
    # Amazon SKIPS; legacy: header_row+1). Writing into that skipped gap
    # silently drops SKUs.
    write_at = _data_start_row(ws, header_row)
    # Clear everything below the field-name row, so the template's prefilled
    # Example row (and its "do not delete this row" instruction row) or a
    # reused workbook's stale rows can't upload as unintended SKUs; the gap
    # rows above write_at are then re-emitted empty by the TSV export.
    if ws.max_row > header_row:
        ws.delete_rows(header_row + 1, ws.max_row - header_row)
    for i, spec_row in enumerate(rows):
        op_token, op_key = _resolve_operation(
            spec_row.get('operation'), schema.dialect
        )
        fields = _row_fields(spec_row, spec, schema)
        # Map any bare/undecorated content field name to this template's
        # actual column (unified decorates them), so the SAME spec fills
        # either dialect. Offer shorthands (our_price/quantity) have no
        # column of their own and pass through to _route_offer_price.
        fields = {schema.resolve_field(k): v for k, v in fields.items()}
        sku = fields.get(schema.field('sku'))
        if not sku:
            raise SystemExit(f'error: row {i} has no sku')

        # Route the child offer price into the marketplace being listed on,
        # so it can't land in the wrong `purchasable_offer` block (which
        # creates an ASIN-only listing stuck in "Missing offer").
        _route_offer_price(fields, cols, mkt_id, i, sku, warnings)

        # Drop an optional Item Highlight the title is too long for, so it
        # can't poison the SKU with a 100476 rejection (SUCCESS OTHER).
        _drop_unusable_item_highlight(fields, i, sku, warnings)

        # Enum validation: reject an invalid operation-family token hard;
        # warn (never fail) on other enums so an unseen-but-valid token
        # from a new category still uploads.
        for fname, fval in fields.items():
            allowed = valid.get(fname)
            if allowed and str(fval).strip().lower() not in allowed:
                warnings.append(
                    f'row {i} sku={sku}: {fname}={fval!r} not in template '
                    f'valid values {sorted(allowed)}'
                )

        # Required-field guard applies to Create rows only. Update/
        # partialupdate touch a subset; delete needs just sku+operation.
        if op_key == 'create':
            parentage = fields.get(schema.field('parentage'), '')
            is_parent = str(parentage).lower() == 'parent'
            no_battery = _is_falsey(
                fields.get('are_batteries_included')
            ) and _is_falsey(fields.get('batteries_required'))
            missing = []
            for f in required:
                if f not in cols or f in fields or f == op_field:
                    continue
                if is_parent and f.startswith(CHILD_LEVEL_PREFIXES):
                    continue
                if no_battery and any(t in f for t in BATTERY_TOKENS):
                    continue
                missing.append(f)
            if missing:
                warnings.append(
                    f'row {i} sku={sku}: missing required field(s) '
                    f'{sorted(missing)}'
                )

        target = ws[write_at + i]
        # Operation column: write the token (blank cell = Create).
        if op_field and op_field in cols:
            target[cols[op_field][0] - 1].value = op_token or None
        for fname, fval in fields.items():
            if fname not in cols:
                unknown_fields.add(fname)
                continue
            # Canonicalise to the template's exact-case enum token when
            # this field has a known valid set -- some fields are
            # case-strict on Amazon's side (e.g. UAE/KSA vs uae/ksa).
            canon = case.get(fname, {}).get(str(fval).strip().lower())
            target[cols[fname][0] - 1].value = canon if canon else fval

    wb.save(args.out)
    # The upload artefact is a tab-delimited .txt, NOT this .xlsm (an
    # openpyxl-saved .xlsm triggers Amazon's 90502 FATAL). Write it next
    # to the .xlsm so the caller uploads the .txt.
    txt_path = args.out.rsplit('.', 1)[0] + '.txt'
    _export_tsv(ws, txt_path)

    for w in warnings:
        print(f'warning: {w}', file=sys.stderr)
    if unknown_fields:
        print(
            f'warning: fields not in this template (skipped): '
            f'{sorted(unknown_fields)}',
            file=sys.stderr,
        )
    # The browser (Ziniao/macOS) can't read /tmp, so uploading a /tmp file
    # silently no-ops (Submit never enables). Nudge the caller to write
    # under the store downloads dir instead.
    if os.path.realpath(txt_path).startswith(('/tmp/', '/private/tmp/')):
        print(
            f'warning: {txt_path} is under /tmp -- the browser cannot read '
            'it and the upload will silently fail. Write --out under '
            '~/.vibe-seller/downloads/<slug>/ instead.',
            file=sys.stderr,
        )
    print(f'wrote {len(rows)} data row(s) -> {args.out}')
    print(
        f'UPLOAD THIS (tab-delimited) -> {txt_path}  '
        '(uploading the Excel file itself triggers a 90502 FATAL)'
    )


def _iter_report_rows(path):
    """Yield rows from a processing report (.xlsm/.xlsx or .txt/.csv)."""
    lower = path.lower()
    if lower.endswith(('.xlsm', '.xlsx')):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for name in wb.sheetnames:
            for row in wb[name].iter_rows(values_only=True):
                yield [('' if c is None else str(c)) for c in row]
        return
    delim = '\t'
    with open(path, encoding='utf-8', errors='replace') as fh:
        sample = fh.readline()
        if ',' in sample and '\t' not in sample:
            delim = ','
        fh.seek(0)
        for line in fh:
            yield line.rstrip('\n').split(delim)


def _template_cell_errors(path):
    """Per-field errors from a processing report's Template tab.

    Amazon writes the precise, field-level fix as CELL COMMENTS (批注) on
    the report's Template tab: the summary table gives code + message,
    but the comment pins the exact column (field) to change. read_only
    mode does not load comments, so open the workbook normally. Yields
    (sku, field_api_name, comment_text) -- the self-correction targets.
    """
    if not path.lower().endswith(('.xlsm', '.xlsx')):
        return
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        if TEMPLATE_SHEET not in wb.sheetnames:
            return
        ws = wb[TEMPLATE_SHEET]
        try:
            header_row = _find_header_row(ws)
        except ValueError:
            return
        names = {
            c.column: str(c.value).strip() for c in ws[header_row] if c.value
        }
        sku_col = next(
            (col for col, n in names.items() if _is_sku_field(n)), None
        )
        for row in ws.iter_rows(min_row=header_row + 1):
            sku = ''
            if sku_col is not None:
                cell = next((c for c in row if c.column == sku_col), None)
                sku = cell.value if cell and cell.value else ''
            for c in row:
                if not (c.comment and c.comment.text):
                    continue
                field = names.get(c.column, f'col{c.column}')
                # A cell comment stacks several messages on Excel's literal
                # `_x000d_` CR marker; split so each is one actionable line.
                raw = str(c.comment.text).replace('_x000d_', '\n')
                for line in raw.split('\n'):
                    line = ' '.join(line.split())
                    if line:
                        yield (str(sku or '?'), field, line)
    finally:
        wb.close()


def _report_comment_errors(path):
    """Per-SKU errors from a processing summary's Template-tab CELL COMMENTS.

    Amazon writes each error/warning as a cell comment on the Template tab
    (orange = error, yellow = warning) keyed to the offending field's
    column -- NOT as a table row. The comment is the only reliable error
    source; a table scan lands on the localized instructions instead.
    Returns [(sku, field, severity, message)].
    """
    if not path.lower().endswith(('.xlsm', '.xlsx')):
        return []
    try:
        wb = openpyxl.load_workbook(path)
    except Exception:
        return []
    if TEMPLATE_SHEET not in wb.sheetnames:
        return []
    ws = wb[TEMPLATE_SHEET]
    try:
        hdr = _find_header_row(ws)
    except ValueError:
        return []
    names = [c.value for c in ws[hdr]]
    sku_col = next(
        (i for i, n in enumerate(names) if n and _is_sku_field(str(n))), None
    )
    if sku_col is None:
        return []
    out = []
    for r in range(hdr + 1, ws.max_row + 1):
        sku = ws.cell(r, sku_col + 1).value
        if not sku:
            continue
        for c in ws[r]:
            if not (c.comment and c.comment.text.strip()):
                continue
            field = names[c.column - 1] if c.column - 1 < len(names) else ''
            text = ' '.join(c.comment.text.split())
            sev = (
                'error'
                if text.lstrip().upper().startswith('ERROR')
                else ('warning' if 'WARNING' in text[:20].upper() else 'info')
            )
            out.append((str(sku), str(field or ''), sev, text[:400]))
    return out


def _write_verdict(batch_id, n_err, n_warn, error_msgs):
    """Write ``BATCH_<id>_VERDICT.json`` to CWD (the task workspace).

    The machine-checkable verdict the completion gate matches against the
    ``UPLOAD_BATCH_<id>.json`` marker bh_upload_flatfile wrote: the task
    cannot finish while a batch has non-image errors. When the caller
    could not extract per-error text, every error counts as non-image
    (conservative -- never lets an unknown error pass as deferrable).
    """
    if not batch_id:
        return
    non_image = [
        m
        for m in error_msgs
        if '18320' not in m and 'main image' not in m.lower()
    ]
    strict = error_msgs or n_err == 0
    with open(f'BATCH_{batch_id}_VERDICT.json', 'w', encoding='utf-8') as fh:
        json.dump(
            {
                'batch_id': batch_id,
                'errors': n_err,
                'warnings': n_warn,
                'non_image_errors': len(non_image) if strict else n_err,
            },
            fh,
        )


def cmd_parse_feedback(args):
    """Summarise Amazon's processing report: per-SKU errors/warnings.

    Prefer the Template-tab cell comments (the authoritative per-field
    error source); fall back to a table scan for report layouts that use
    one. A parent SKU's errors block its children -- fix the parent first.
    """
    comment_errs = _report_comment_errors(args.file)
    if comment_errs:
        n_err = sum(1 for _s in comment_errs if _s[2] == 'error')
        n_warn = sum(1 for _s in comment_errs if _s[2] == 'warning')
        for sku, field, sev, msg in comment_errs:
            print(f'  [{sev}] sku={sku} field={field}: {msg}')
        print(
            f'\n{n_err} error(s), {n_warn} warning(s) across '
            f'{len({s for s, *_ in comment_errs})} SKU(s).'
        )
        _write_verdict(
            getattr(args, 'batch_id', None),
            n_err,
            n_warn,
            [m for _s, _f, sev, m in comment_errs if sev == 'error'],
        )
        if n_err:
            print(
                'NOT DONE. Fix ALL errors (parent first) and re-upload. A '
                'SKU with any error is either not created OR created but '
                'flagged "Action required" (SUCCESS OTHER) -- it shows up in '
                'inventory yet the error is UNRESOLVED. Inventory presence '
                'is NOT "done"; only 18320 (missing main image) is a legit '
                'deferral. Re-run parse-feedback on the new report until the '
                'sole remaining error is 18320.'
            )
            # Non-zero exit so a caller/CI sees the feed had blocking
            # errors -- matches the table-scan path below.
            sys.exit(1)
        return

    rows = list(_iter_report_rows(args.file))
    if not rows:
        raise SystemExit('error: empty report')

    def idx_of(header, *needles):
        for j, cell in enumerate(header):
            c = cell.strip().lower()
            if any(n in c for n in needles):
                return j
        return None

    hdr_i = None
    for i, row in enumerate(rows):
        joined = ' '.join(row).lower()
        if ('error' in joined or 'sku' in joined) and (
            'sku' in joined or 'record' in joined
        ):
            hdr_i = i
            break

    n_err = n_warn = 0
    if hdr_i is not None:
        header = rows[hdr_i]
        c_sku = idx_of(header, 'sku')
        c_type = idx_of(header, 'error type', 'type', 'severity')
        c_msg = idx_of(header, 'message', 'error message', 'description')
        c_code = idx_of(header, 'code')
        print(f'report table header at row {hdr_i}: {header}')
        for row in rows[hdr_i + 1 :]:
            if not any(cell.strip() for cell in row):
                continue
            sku = row[c_sku] if c_sku is not None and c_sku < len(row) else ''
            typ = (
                row[c_type] if c_type is not None and c_type < len(row) else ''
            )
            msg = row[c_msg] if c_msg is not None and c_msg < len(row) else ''
            code = (
                row[c_code] if c_code is not None and c_code < len(row) else ''
            )
            tl = typ.strip().lower()
            if tl == 'error' or 'error' in tl:
                n_err += 1
            elif 'warn' in tl:
                n_warn += 1
            print(f'  [{typ or "?"}] sku={sku} code={code}: {msg}')
    else:
        print('(no structured table found; rows mentioning "error")')
        for row in rows:
            if 'error' in ' '.join(row).lower():
                print('  ' + ' | '.join(row))
                n_err += 1

    # The precise, field-level fixes live in the Template-tab cell
    # comments (批注). Print them as (sku, field, message) -- for each,
    # set that field to a valid value (see `inspect`) and re-upload. The
    # error set is category-specific and unbounded, so extract it from
    # the report rather than hardcode fixes.
    cell_errs = list(_template_cell_errors(args.file))
    if cell_errs:
        print(
            '\nper-field cell comments (批注) -- fix the exact field each names:'
        )
        for sku, field, msg in cell_errs:
            print(f'  sku={sku} field={field}: {msg}')
        # The 批注 are the authoritative per-field verdict; the summary
        # table above is often mislocated, so count from these instead.
        c_err = sum(
            1 for _, _, m in cell_errs if m.lstrip().upper().startswith('ERROR')
        )
        c_warn = sum(
            1 for _, _, m in cell_errs if m.lstrip().upper().startswith('WARN')
        )
        if c_err or c_warn:
            n_err, n_warn = c_err, c_warn

    print(f'\nsummary: {n_err} error(s), {n_warn} warning(s)')
    _write_verdict(
        getattr(args, 'batch_id', None),
        n_err,
        n_warn,
        [m for _s, _f, m in cell_errs] if cell_errs else [],
    )
    if n_err:
        sys.exit(1)


def main():
    ap = argparse.ArgumentParser(description=__doc__.split('\n')[0])
    sub = ap.add_subparsers(dest='cmd', required=True)

    p = sub.add_parser('inspect', help='dump fields / required / enums')
    p.add_argument('file')
    p.add_argument('--field', help='detail for one field API name')
    p.set_defaults(func=cmd_inspect)

    p = sub.add_parser('fill', help='write parent/child rows from a spec')
    p.add_argument('file', help='the category template (.xlsm)')
    p.add_argument('--spec', required=True, help='JSON spec of rows')
    p.add_argument('--out', required=True, help='output .xlsm path')
    p.add_argument(
        '--marketplace',
        help='country code (SA/AE/AU/…) or raw marketplace id you are '
        'listing on; routes the offer price to the right block. Overrides '
        'the spec\'s top-level "marketplace".',
    )
    p.set_defaults(func=cmd_fill)

    p = sub.add_parser('parse-feedback', help='summarise a processing report')
    p.add_argument('file')
    p.add_argument(
        '--batch-id',
        help='write BATCH_<id>_VERDICT.json for the completion gate',
    )
    p.set_defaults(func=cmd_parse_feedback)

    args = ap.parse_args()
    args.func(args)


if __name__ == '__main__':
    main()
