#!/usr/bin/env python3
"""Taobao/Tmall 千牛 商品 Excel bulk round-trip helper.

Drives the 商品管理 → 更多批量操作 → excel商品批量导出 / excel商品批量编辑 /
批量导入 loop: inspect an exported workbook, fill edits keyed by the
workbook's OWN column headers, and summarise the import result.

Why header-keyed (vs positional)
--------------------------------
The 商品 export's column set is category-specific and changes over time,
and the header labels are localised Chinese. Keying every value by the
export's own header string (discovered with `inspect`) keeps the tool
robust across categories and template revisions -- the same philosophy
as the amazon-listing flat-file tool. `fill` copies the exported
workbook verbatim and only overwrites the named cells, so an import
changes exactly what you intended and nothing else.

Usage
-----
  listing_bulk.py inspect       WORKBOOK.xlsx
  listing_bulk.py fill          WORKBOOK.xlsx --spec SPEC.json --out OUT.xlsx
  listing_bulk.py parse-import  RESULT.xlsx

Requires: openpyxl.
"""

import argparse
import json
import shutil
import sys

try:
    import openpyxl
except ImportError:
    sys.exit('error: openpyxl is required (pip install openpyxl)')


# Tokens that identify the header row and the per-商品 identity column.
# Taobao/Tmall exports label the identity column one of these.
IDENTITY_TOKENS = ('商品ID', '宝贝ID', '商品编码', '数字ID', 'itemId', '商品id')
# Substrings that mark a plausible header row (any one is enough).
HEADER_TOKENS = IDENTITY_TOKENS + ('商品标题', '标题', '价格', '一口价', '库存')


def _find_header_row(ws, max_scan=12):
    """Return (row_index_1based, [header strings]) for the header row.

    Picks the first of the first `max_scan` rows that contains a known
    header token; falls back to the row with the most non-empty string
    cells. Taobao exports sometimes carry a title/notice row above the
    real header, so we scan rather than assume row 1.
    """
    best = None
    for r in range(1, max_scan + 1):
        vals = [
            ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)
        ]
        strs = [str(v).strip() for v in vals if v not in (None, '')]
        joined = ' '.join(strs)
        if any(tok in joined for tok in HEADER_TOKENS):
            return r, [('' if v is None else str(v).strip()) for v in vals]
        # track the densest row as a fallback
        if best is None or len(strs) > best[0]:
            best = (
                len(strs),
                r,
                [('' if v is None else str(v).strip()) for v in vals],
            )
    if best:
        return best[1], best[2]
    raise SystemExit('error: could not locate a header row in the workbook')


def _identity_col(headers):
    """1-based column index of the 商品 identity column, or None."""
    for tok in IDENTITY_TOKENS:
        for i, h in enumerate(headers):
            if h and tok in h:
                return i + 1
    return None


def _sheet(wb):
    return wb[wb.sheetnames[0]]


def cmd_inspect(args):
    wb = openpyxl.load_workbook(args.file, read_only=True, data_only=True)
    ws = _sheet(wb)
    header_row, headers = _find_header_row(ws)
    cols = [h for h in headers if h]
    id_col = _identity_col(headers)
    id_name = headers[id_col - 1] if id_col else None
    n_data = sum(
        1
        for r in range(header_row + 1, ws.max_row + 1)
        if any(
            ws.cell(row=r, column=c).value not in (None, '')
            for c in range(1, ws.max_column + 1)
        )
    )
    print(f'sheet        : {wb.sheetnames[0]}')
    print(f'header row   : {header_row} (data starts row {header_row + 1})')
    print(
        f'identity col : {id_name or "<<none found — check IDENTITY_TOKENS>>"}'
    )
    print(f'data rows    : {n_data}')
    print(f'columns ({len(cols)}):')
    for h in cols:
        print(f'  {h}')
    wb.close()


def _load_spec(path):
    with open(path, encoding='utf-8') as fh:
        spec = json.load(fh)
    rows = spec.get('rows') or []
    if not rows:
        raise SystemExit('error: spec has no rows')
    return rows


def cmd_fill(args):
    rows = _load_spec(args.spec)
    shutil.copyfile(args.file, args.out)
    wb = openpyxl.load_workbook(args.out)  # keep styles/other cells verbatim
    ws = _sheet(wb)
    header_row, headers = _find_header_row(ws)
    col_of = {h: i + 1 for i, h in enumerate(headers) if h}
    id_col = _identity_col(headers)
    if not id_col:
        raise SystemExit(
            'error: no identity column (商品ID/宝贝ID/…) — cannot match rows'
        )
    id_name = headers[id_col - 1]

    # Map identity value -> worksheet row.
    row_of = {}
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=id_col).value
        if v not in (None, ''):
            row_of[str(v).strip()] = r

    warnings, written = [], 0
    for i, spec_row in enumerate(rows):
        ident = str(
            spec_row.get('商品ID') or spec_row.get(id_name) or ''
        ).strip()
        if not ident:
            raise SystemExit(f'error: spec row {i} has no 商品ID')
        r = row_of.get(ident)
        if not r:
            warnings.append(f'row {i}: {id_name}={ident} not in the export')
            continue
        for col_name, val in (spec_row.get('fields') or {}).items():
            if col_name not in col_of:
                warnings.append(
                    f'row {i} ({ident}): column {col_name!r} not in the export'
                )
                continue
            ws.cell(row=r, column=col_of[col_name]).value = val
            written += 1

    wb.save(args.out)
    for w in warnings:
        print(f'warning: {w}', file=sys.stderr)
    print(
        f'wrote {written} cell(s) across {len(rows)} spec row(s) -> {args.out}'
    )
    print('IMPORT IS A WRITE — have the user review the diff before 批量导入.')


def cmd_parse_import(args):
    """Summarise the import-result workbook: per-row 成功/失败 + reason."""
    wb = openpyxl.load_workbook(args.file, read_only=True, data_only=True)
    ws = _sheet(wb)
    header_row, headers = _find_header_row(ws)

    def find(*needles):
        for i, h in enumerate(headers):
            if h and any(n in h for n in needles):
                return i + 1
        return None

    c_id = _identity_col(headers)
    c_status = find('结果', '状态', '成功', '是否成功')
    c_reason = find('原因', '失败原因', '错误', '备注', 'message')

    def cell(r, c):
        return str(ws.cell(row=r, column=c).value or '').strip() if c else ''

    n_ok = n_fail = 0
    for r in range(header_row + 1, ws.max_row + 1):
        ident = cell(r, c_id)
        status = cell(r, c_status)
        reason = cell(r, c_reason)
        if not (ident or status or reason):
            continue
        fail = (
            ('失败' in status)
            or ('错误' in status)
            or bool(reason and '成功' not in status)
        )
        if fail:
            n_fail += 1
        else:
            n_ok += 1
        print(f'  [{status or "?"}] {ident}{": " + reason if reason else ""}')
    print(f'\nsummary: {n_ok} ok, {n_fail} failed')
    if n_fail:
        sys.exit(1)


def main():
    ap = argparse.ArgumentParser(description=__doc__.split('\n')[0])
    sub = ap.add_subparsers(dest='cmd', required=True)
    p = sub.add_parser('inspect', help='dump header row + columns + row count')
    p.add_argument('file')
    p.set_defaults(func=cmd_inspect)
    p = sub.add_parser('fill', help='set cells by header name, keyed by 商品ID')
    p.add_argument('file')
    p.add_argument('--spec', required=True)
    p.add_argument('--out', required=True)
    p.set_defaults(func=cmd_fill)
    p = sub.add_parser(
        'parse-import', help='summarise an import-result workbook'
    )
    p.add_argument('file')
    p.set_defaults(func=cmd_parse_import)
    args = ap.parse_args()
    args.func(args)


if __name__ == '__main__':
    main()
