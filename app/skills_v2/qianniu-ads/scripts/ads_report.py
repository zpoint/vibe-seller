#!/usr/bin/env python3
"""Parse a 万相台无界 报表 export into normalised TSV.

Reads a 计划/商品 report exported from one.alimama.com (a GB18030 CSV or an
`.xlsx`), locates the header row by its own Chinese labels, and emits a
normalised, tab-separated table (stable English keys) to stdout for the
audit. Header-keyed, so it survives column reordering / new columns.

Usage
-----
  ads_report.py REPORT.csv            # or REPORT.xlsx
  ads_report.py REPORT.csv --raw      # dump the detected headers + first rows

Note: Taobao/Tmall CSV exports are GB18030, not UTF-8 (see qianniu-shared).
Requires: openpyxl (only for .xlsx).
"""

import argparse
import csv
import sys

try:
    import openpyxl
except ImportError:  # only needed for .xlsx reports
    openpyxl = None

# Normalised key <- any header substring that denotes it. First match wins.
FIELD_MAP = [
    ('campaign_id', ('计划ID', '计划 ID')),
    ('campaign', ('计划名', '计划名称', '计划')),
    ('product_id', ('宝贝ID', '商品ID')),
    ('spend', ('花费', '消耗')),
    ('gmv', ('成交金额', '总成交金额', '净成交金额')),
    ('roi', ('投产比', 'ROI', 'roi')),
    ('orders', ('成交笔数', '总成交笔数')),
    ('clicks', ('点击量', '点击数')),
    ('ctr', ('点击率',)),
    ('impressions', ('展现量', '展现数')),
    ('cpc', ('平均点击花费', '点击花费')),
]
HEADER_TOKENS = ('花费', '成交', '计划', '宝贝', '点击', '展现', '投产比')


def _rows_from_file(path):
    """Yield the raw rows (list of str) from a CSV (gb18030) or xlsx."""
    if path.lower().endswith(('.xlsx', '.xlsm')):
        if openpyxl is None:
            raise SystemExit(
                'error: openpyxl is required for .xlsx (pip install openpyxl)'
            )
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(values_only=True):
            yield [('' if c is None else str(c)) for c in row]
        wb.close()
        return
    # CSV — Taobao exports are GB18030; fall back to utf-8-sig.
    for enc in ('gb18030', 'utf-8-sig'):
        try:
            with open(path, encoding=enc, newline='') as fh:
                yield from csv.reader(fh)
            return
        except UnicodeDecodeError:
            continue
    raise SystemExit('error: could not decode CSV as GB18030 or UTF-8')


def _find_header(rows, max_scan=8):
    for i, r in enumerate(rows[:max_scan]):
        if sum(any(tok in c for tok in HEADER_TOKENS) for c in r) >= 2:
            return i
    raise SystemExit('error: could not locate the report header row')


def _col_index(headers):
    """Map normalised key -> column index (first header substring match)."""
    out = {}
    for key, needles in FIELD_MAP:
        for i, h in enumerate(headers):
            if any(n in h for n in needles):
                out[key] = i
                break
    return out


def main():
    ap = argparse.ArgumentParser(description=__doc__.split('\n')[0])
    ap.add_argument('file')
    ap.add_argument('--raw', action='store_true', help='dump headers + sample')
    args = ap.parse_args()

    rows = list(_rows_from_file(args.file))
    if not rows:
        raise SystemExit('error: empty report')
    hi = _find_header(rows)
    headers = [c.strip() for c in rows[hi]]
    if args.raw:
        print('header row', hi, ':', ' | '.join(h for h in headers if h))
        for r in rows[hi + 1 : hi + 4]:
            print('  ', ' | '.join(r))
        return

    idx = _col_index(headers)
    keys = [k for k, _ in FIELD_MAP if k in idx]
    if not keys:
        raise SystemExit(
            'error: no known metric columns found — run with --raw to see the '
            'actual headers and extend FIELD_MAP'
        )
    w = csv.writer(sys.stdout, delimiter='\t', lineterminator='\n')
    w.writerow(keys)
    n = 0
    for r in rows[hi + 1 :]:
        if not any(c.strip() for c in r):
            continue
        vals = [r[idx[k]].strip() if idx[k] < len(r) else '' for k in keys]
        if not any(vals):
            continue
        w.writerow(vals)
        n += 1
    print(f'# {n} rows, columns: {", ".join(keys)}', file=sys.stderr)


if __name__ == '__main__':
    main()
