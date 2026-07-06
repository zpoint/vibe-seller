#!/usr/bin/env python3
"""Amazon Ads bulk-sheet analysis + edit tool (Sponsored Products).

Reads a `Download campaigns` bulk export (XLSX) and either summarises it
(for an ad-tuning report) or emits an upload-ready sheet that CREATEs a
new manual-keyword campaign or UPDATEs bids on existing keywords.

Why this exists
---------------
The click-through create/tune path is slow and brittle (kat-* shadow
components, hover menus, per-field modals). Amazon's own Bulk Operations
export/import is the supported batch API: download the campaigns sheet,
edit rows, re-upload. This script is the "edit rows" step.

Locale-generality (this is load-bearing)
----------------------------------------
The Sponsored Products sheet has a FIXED 52-column order that is
identical in every console language; only the sheet *name* and the
*header text* are localised. So:

  * Columns are addressed by POSITION (see `Col`), never by header text.
  * The header row is validated against a bilingual dictionary and a
    warning is printed on mismatch, but parsing never depends on it.
  * When emitting an upload sheet we CLONE the downloaded workbook's
    header row and sheet name verbatim, then append data rows. Amazon's
    upload validator reads that header row, so reusing the exact one the
    account produced is the only header we know it will accept -- no
    guessing at localised header strings.

Entity rows are matched by their position in the enum, resolved the same
way (English name, else any known localised name). Add new localisations
to `SHEET_NAMES` / `ENTITY` as they are observed.

Safety
------
  * New campaigns are emitted `paused` with a caller-supplied daily
    budget (default minimal) so nothing spends on upload.
  * Product Ad rows are guarded against the ASIN-as-SKU trap: an ASIN in
    the SKU field makes Amazon silently drop the whole ad group. If the
    SKU looks like an ASIN the script refuses (see `looks_like_asin`).

Usage
-----
  ads_bulk.py inspect        EXPORT.xlsx
  ads_bulk.py clone-campaign EXPORT.xlsx --src "<name>" --new "<name>" \
                             --sku SELLER-SKU --asin B0XXXXXXXX \
                             [--daily-budget 1] [--default-bid 0.75] \
                             [--out OUT.xlsx]
  ads_bulk.py bid-update     EXPORT.xlsx --campaign "<name>" \
                             (--scale 0.85 | --set-bid 1.20) \
                             [--out OUT.xlsx]

Requires: openpyxl.
"""

import argparse
import datetime
import json
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit('error: openpyxl is required (pip install openpyxl)')


# --- 52-column schema, addressed positionally (§ mechanics 4c) ---------
class Col:
    PRODUCT = 0
    ENTITY = 1
    OPERATION = 2
    CAMPAIGN_ID = 3
    AD_GROUP_ID = 4
    PORTFOLIO_ID = 5
    AD_ID = 6
    KEYWORD_ID = 7
    PRODUCT_TARGETING_ID = 8
    CAMPAIGN_NAME = 9
    AD_GROUP_NAME = 10
    CAMPAIGN_NAME_INFO = 11
    AD_GROUP_NAME_INFO = 12
    PORTFOLIO_NAME_INFO = 13
    START_DATE = 14
    END_DATE = 15
    TARGETING_TYPE = 16
    STATE = 17
    CAMPAIGN_STATE_INFO = 18
    AD_GROUP_STATE_INFO = 19
    DAILY_BUDGET = 20
    SKU = 21
    ASIN_INFO = 22
    ELIGIBILITY_INFO = 23
    INELIGIBILITY_REASON_INFO = 24
    AD_GROUP_DEFAULT_BID = 25
    AD_GROUP_DEFAULT_BID_INFO = 26
    BID = 27
    KEYWORD_TEXT = 28
    NATIVE_LANGUAGE_KEYWORD = 29
    NATIVE_LANGUAGE_LOCALE = 30
    MATCH_TYPE = 31
    BIDDING_STRATEGY = 32
    PLACEMENT = 33
    PERCENTAGE = 34
    PRODUCT_TARGETING_EXPR = 35
    RESOLVED_TARGETING_EXPR_INFO = 36
    AUDIENCE_ID = 37
    SHOPPER_COHORT_PCT = 38
    SHOPPER_COHORT_TYPE = 39
    SEGMENT_NAME_INFO = 40
    IMPRESSIONS = 41
    CLICKS = 42
    CTR = 43
    SPEND = 44
    SALES = 45
    ORDERS = 46
    UNITS = 47
    CONVERSION_RATE = 48
    ACOS = 49
    CPC = 50
    ROAS = 51


NUM_COLS = 52

# Sheet names by locale (English first). Fallback: the 52-column sheet.
SHEET_NAMES = ('Sponsored Products Campaigns', '商品推广活动')

# Entity enum values by locale (add localisations as observed).
ENTITY = {
    'campaign': ('Campaign', '广告活动'),
    'bidding_adjustment': ('Bidding Adjustment', '竞价调整'),
    'ad_group': ('Ad Group', '广告组'),
    'product_ad': ('Product Ad', '商品广告'),
    'keyword': ('Keyword', '关键词'),
    'negative_keyword': ('Negative Keyword', '否定关键词'),
    'campaign_negative_keyword': (
        'Campaign Negative Keyword',
        '广告活动否定关键词',
    ),
    'product_targeting': ('Product Targeting', '商品定向'),
}

# IMPORTANT — upload uses ENGLISH API tokens, NOT the localised strings
# the export DISPLAYS. The export's own `Config` sheet lists the valid
# upload values (SponsoredProducts*States/OperationNames/TargetingTypes/
# MatchTypes/Strategys) and they are all English. So a keyword the export
# shows as `广泛` / `精准` must be written back as `broad` / `exact`, or
# Amazon rejects the row. Verified live: a zh_CN account rejected an
# upload carrying localised match tokens with "0 of N uploaded".
STATE_ENABLED = 'enabled'
STATE_PAUSED = 'paused'

# Config: SponsoredProductsCreateCampaignTargetingTypes = AUTO | MANUAL.
TARGETING_MANUAL = 'MANUAL'
TARGETING_AUTO = 'AUTO'

# Config: SponsoredProductsCreateKeywordMatchTypes = exact|phrase|broad.
# Map the localised DISPLAY tokens seen in exports -> the API token.
MATCH_TYPE_API = {
    'broad': 'broad',
    'phrase': 'phrase',
    'exact': 'exact',
    '广泛': 'broad',
    '词组': 'phrase',
    '精准': 'exact',
}


def match_type_api(display):
    """Normalise an export's (possibly localised) match type to the API
    token Amazon's uploader accepts. Falls back to a lowercased value."""
    if display is None:
        return None
    key = str(display).strip()
    return MATCH_TYPE_API.get(key, key.lower())


# Config: *States = enabled|paused|archived. Map localised display -> API.
STATE_API = {
    'enabled': 'enabled',
    'paused': 'paused',
    'archived': 'archived',
    '已启用': 'enabled',
    '已暂停': 'paused',
    '已归档': 'archived',
}


def state_api(display):
    """Normalise an export's (possibly localised) state to the API token."""
    if display is None:
        return None
    return STATE_API.get(str(display).strip(), str(display).strip().lower())


# Header row validation dictionary (position -> accepted labels). Only a
# few load-bearing columns; parsing never depends on this.
HEADER_CHECK = {
    Col.ENTITY: ('Entity', '实体层级'),
    Col.OPERATION: ('Operation', '操作'),
    Col.CAMPAIGN_NAME: ('Campaign Name', '广告活动名称'),
    Col.SKU: ('SKU',),
    Col.BID: ('Bid', '竞价'),
    Col.MATCH_TYPE: ('Match Type', '匹配类型'),
}

ASIN_RE = re.compile(r'^B0[0-9A-Z]{8}$')


def looks_like_asin(value):
    """True if value matches an Amazon ASIN (B0 + 8 alphanumerics)."""
    return bool(ASIN_RE.match(str(value or '').strip().upper()))


def find_sp_sheet(wb):
    """Return the Sponsored Products Campaigns worksheet, locale-safe."""
    for name in SHEET_NAMES:
        if name in wb.sheetnames:
            return wb[name]
    # Fallback: the sheet whose header row is 52 wide.
    for ws in wb.worksheets:
        if ws.max_column == NUM_COLS and ws.max_row > 1:
            return ws
    sys.exit(
        'error: could not locate the Sponsored Products sheet '
        '(known names: {}, none 52-col wide)'.format(', '.join(SHEET_NAMES))
    )


def validate_header(header):
    """Warn (do not fail) if the positional header looks unexpected."""
    if len(header) < NUM_COLS:
        print(
            f'warning: header has {len(header)} cols, expected '
            f'{NUM_COLS} -- positional mapping may be wrong',
            file=sys.stderr,
        )
    for idx, accepted in HEADER_CHECK.items():
        got = str(header[idx]).strip() if idx < len(header) else ''
        if got not in accepted:
            print(
                f'warning: col {idx} header {got!r} not in known set '
                f'{accepted} (new locale? parsing still positional)',
                file=sys.stderr,
            )


def load(path):
    """Load workbook + SP sheet rows (header, data_rows)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = find_sp_sheet(wb)
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit(f'error: SP sheet {ws.title!r} is empty')
    header, data = rows[0], [list(r) for r in rows[1:]]
    validate_header(header)
    return wb, ws, header, data


def is_entity(row, kind):
    """True if row's Entity cell is any localised name for `kind`."""
    return str(row[Col.ENTITY]) in ENTITY[kind]


def campaign_name_of(row, id_to_name):
    """Resolve a row's campaign name (child rows use the info column)."""
    return (
        row[Col.CAMPAIGN_NAME]
        or row[Col.CAMPAIGN_NAME_INFO]
        or id_to_name.get(row[Col.CAMPAIGN_ID])
    )


def build_campaign_index(data):
    """Map campaign id -> campaign name from Campaign rows."""
    return {
        r[Col.CAMPAIGN_ID]: r[Col.CAMPAIGN_NAME]
        for r in data
        if is_entity(r, 'campaign')
    }


# --- inspect ------------------------------------------------------------
def cmd_inspect(args):
    _wb, ws, _header, data = load(args.file)
    print(f'sheet: {ws.title}   rows: {len(data)}')
    counts = {}
    for r in data:
        counts[str(r[Col.ENTITY])] = counts.get(str(r[Col.ENTITY]), 0) + 1
    print('\nentity counts:')
    for k, v in sorted(counts.items(), key=lambda kv: -kv[1]):
        print(f'  {k:<24} {v}')

    id_to_name = build_campaign_index(data)
    print('\ncampaigns:')
    for r in data:
        if not is_entity(r, 'campaign'):
            continue
        print(
            f'  {r[Col.CAMPAIGN_NAME]:<45} state={r[Col.STATE]!s:<10} '
            f'budget={r[Col.DAILY_BUDGET]!s:<6} spend={r[Col.SPEND]!s:<8} '
            f'sales={r[Col.SALES]!s:<8} acos={r[Col.ACOS]}'
        )

    # SKU naming scheme (Product Ad rows) -- the §4d canonical source.
    print('\nProduct Ad SKUs (Entity=Product Ad):')
    seen = set()
    for r in data:
        if is_entity(r, 'product_ad') and r[Col.SKU] not in seen:
            seen.add(r[Col.SKU])
            print(
                f'  camp={campaign_name_of(r, id_to_name)!s:<40} '
                f'SKU={r[Col.SKU]!s:<24} ASIN={r[Col.ASIN_INFO]}'
            )


# --- shared upload-sheet writer ----------------------------------------
def blank_row():
    return [None] * NUM_COLS


def write_upload(ws_title, header, out_rows, out_path):
    """Write an upload workbook: cloned header + Create/Update data rows.

    Reusing the downloaded header row verbatim is what makes the upload
    locale-robust -- it is the exact header this account's Amazon emits
    and therefore the only one we know its validator accepts.
    """
    out = openpyxl.Workbook()
    ws = out.active
    ws.title = ws_title
    ws.append(list(header))
    for r in out_rows:
        ws.append(r)
    out.save(out_path)
    print(f'wrote {len(out_rows)} data row(s) -> {out_path}')


# --- clone-campaign -----------------------------------------------------
def cmd_clone_campaign(args):
    if looks_like_asin(args.sku):
        sys.exit(
            f'error: --sku {args.sku!r} looks like an ASIN. The Product Ad SKU '
            'must be the real seller SKU, or Amazon silently drops '
            'the entire ad group (mechanics §4b). Pass the seller '
            'SKU; use --asin for the ASIN column.'
        )

    _wb, ws, header, data = load(args.file)
    id_to_name = build_campaign_index(data)

    # Collect the source campaign's keyword rows.
    src_keywords = []
    for r in data:
        if not is_entity(r, 'keyword'):
            continue
        if campaign_name_of(r, id_to_name) == args.src:
            src_keywords.append(r)
    if not src_keywords:
        sys.exit(
            f'error: no keyword rows found for source campaign {args.src!r} '
            '(check the exact name via `inspect`)'
        )

    new = args.new
    ag = args.ad_group or new
    paused = STATE_PAUSED  # English API token (Config: enabled|paused)
    # Start Date is REQUIRED for a Create Campaign row (Config:
    # SponsoredProductsCreateCampaignRequiredHeaders). Format YYYYMMDD.
    start_date = args.start_date or datetime.date.today().strftime('%Y%m%d')
    # Create rows require Campaign Id / Ad Group Id (Config
    # *CreateRequiredHeaders). For NEW entities these are PLACEHOLDER
    # ids: Amazon assigns real ids on creation and uses these only to
    # link parent->child rows within the sheet. Any unique string works;
    # we reuse the names so the linkage is human-readable.
    camp_id = new
    ag_id = ag
    out_rows = []

    # Campaign
    c = blank_row()
    c[Col.PRODUCT] = 'Sponsored Products'
    c[Col.ENTITY] = ENTITY['campaign'][0]
    c[Col.OPERATION] = 'Create'
    c[Col.CAMPAIGN_ID] = camp_id
    c[Col.CAMPAIGN_NAME] = new
    c[Col.START_DATE] = start_date
    c[Col.TARGETING_TYPE] = TARGETING_MANUAL
    c[Col.STATE] = paused
    c[Col.DAILY_BUDGET] = args.daily_budget
    c[Col.BIDDING_STRATEGY] = args.bidding_strategy
    out_rows.append(c)

    # Ad Group
    g = blank_row()
    g[Col.PRODUCT] = 'Sponsored Products'
    g[Col.ENTITY] = ENTITY['ad_group'][0]
    g[Col.OPERATION] = 'Create'
    g[Col.CAMPAIGN_ID] = camp_id
    g[Col.AD_GROUP_ID] = ag_id
    g[Col.CAMPAIGN_NAME] = new
    g[Col.AD_GROUP_NAME] = ag
    g[Col.STATE] = paused
    g[Col.AD_GROUP_DEFAULT_BID] = args.default_bid
    out_rows.append(g)

    # Product Ad (SKU guarded above)
    p = blank_row()
    p[Col.PRODUCT] = 'Sponsored Products'
    p[Col.ENTITY] = ENTITY['product_ad'][0]
    p[Col.OPERATION] = 'Create'
    p[Col.CAMPAIGN_ID] = camp_id
    p[Col.AD_GROUP_ID] = ag_id
    p[Col.CAMPAIGN_NAME] = new
    p[Col.AD_GROUP_NAME] = ag
    p[Col.STATE] = paused
    p[Col.SKU] = args.sku
    # ASIN column is informational (the uploader keys on SKU), but write
    # it when given so the generated sheet is self-documenting / verifiable.
    if args.asin:
        p[Col.ASIN_INFO] = args.asin
    out_rows.append(p)

    # Keywords, copied from source (text + match type; bid = source bid
    # unless overridden). Native-language columns are copied through so
    # non-Latin keyword text survives the round trip.
    for r in src_keywords:
        k = blank_row()
        k[Col.PRODUCT] = 'Sponsored Products'
        k[Col.ENTITY] = ENTITY['keyword'][0]
        k[Col.OPERATION] = 'Create'
        k[Col.CAMPAIGN_ID] = camp_id
        k[Col.AD_GROUP_ID] = ag_id
        k[Col.CAMPAIGN_NAME] = new
        k[Col.AD_GROUP_NAME] = ag
        k[Col.STATE] = paused
        k[Col.BID] = args.keyword_bid or r[Col.BID]
        k[Col.KEYWORD_TEXT] = r[Col.KEYWORD_TEXT]
        k[Col.NATIVE_LANGUAGE_KEYWORD] = r[Col.NATIVE_LANGUAGE_KEYWORD]
        k[Col.NATIVE_LANGUAGE_LOCALE] = r[Col.NATIVE_LANGUAGE_LOCALE]
        # Normalise the export's DISPLAY match token (e.g. 广泛) to the
        # English API token (broad) the uploader requires.
        k[Col.MATCH_TYPE] = match_type_api(r[Col.MATCH_TYPE])
        out_rows.append(k)

    print(
        f'cloning {len(src_keywords)} keyword(s) from {args.src!r} into '
        f'new campaign {new!r} (paused, budget={args.daily_budget}, '
        f'sku={args.sku})'
    )
    out_path = args.out or 'bulk_create_upload.xlsx'
    write_upload(ws.title, header, out_rows, out_path)


# --- bid-update ---------------------------------------------------------
def cmd_bid_update(args):
    if args.scale is None and args.set_bid is None:
        sys.exit('error: pass --scale FACTOR or --set-bid VALUE')
    _wb, ws, header, data = load(args.file)
    id_to_name = build_campaign_index(data)

    out_rows = []
    for r in data:
        if not is_entity(r, 'keyword'):
            continue
        if args.campaign and campaign_name_of(r, id_to_name) != args.campaign:
            continue
        cur = r[Col.BID]
        if cur in (None, ''):
            continue
        if args.set_bid is not None:
            new_bid = args.set_bid
        else:
            new_bid = round(float(cur) * args.scale, 2)
        if float(new_bid) == float(cur):
            continue
        u = blank_row()
        u[Col.PRODUCT] = 'Sponsored Products'
        u[Col.ENTITY] = ENTITY['keyword'][0]
        u[Col.OPERATION] = 'Update'
        # Identify the existing keyword by its ids (preferred) + text.
        u[Col.CAMPAIGN_ID] = r[Col.CAMPAIGN_ID]
        u[Col.AD_GROUP_ID] = r[Col.AD_GROUP_ID]
        u[Col.KEYWORD_ID] = r[Col.KEYWORD_ID]
        u[Col.CAMPAIGN_NAME] = campaign_name_of(r, id_to_name)
        u[Col.KEYWORD_TEXT] = r[Col.KEYWORD_TEXT]
        # State is REQUIRED for an Update Keyword row (Config); preserve
        # the keyword's current state as the API token. Match type
        # normalised to the API token too.
        u[Col.STATE] = state_api(r[Col.STATE]) or STATE_ENABLED
        u[Col.MATCH_TYPE] = match_type_api(r[Col.MATCH_TYPE])
        u[Col.BID] = new_bid
        out_rows.append(u)

    if not out_rows:
        sys.exit('error: no keyword bids matched / changed')
    how = (
        f'scale x{args.scale}'
        if args.scale is not None
        else f'set={args.set_bid}'
    )
    print(f'bid update: {len(out_rows)} keyword row(s) {how}')
    out_path = args.out or 'bulk_bid_update.xlsx'
    write_upload(ws.title, header, out_rows, out_path)


# --- archive-campaign ---------------------------------------------------
def cmd_archive_campaign(args):
    """Emit an Archive row for a campaign (Config: Archive Campaign needs
    only the real Campaign Id). Archiving the campaign cascades to its ad
    groups / ads / keywords. Use this to disable + clean up a campaign.

    The campaign must exist in the export, so its real Campaign Id can be
    read. A paused/zero-impression campaign only appears when the export
    was taken with the zero-impression box checked (see bulk-operations
    §download)."""
    _wb, ws, header, data = load(args.file)
    cid = None
    for r in data:
        if is_entity(r, 'campaign') and r[Col.CAMPAIGN_NAME] == args.campaign:
            cid = r[Col.CAMPAIGN_ID]
            break
    if not cid:
        sys.exit(
            f'error: campaign {args.campaign!r} not found in the export '
            '(a paused/zero-impression campaign needs a zero-impression '
            'export to appear). Check the exact name via `inspect`.'
        )
    a = blank_row()
    a[Col.PRODUCT] = 'Sponsored Products'
    a[Col.ENTITY] = ENTITY['campaign'][0]
    a[Col.OPERATION] = 'Archive'
    a[Col.CAMPAIGN_ID] = cid
    print(f'archiving campaign {args.campaign!r} (id={cid})')
    out_path = args.out or 'bulk_archive.xlsx'
    write_upload(ws.title, header, [a], out_path)


def cmd_scope(args):
    """Print the ACTIVE (state=enabled) Campaign ids for AUDIT_SCOPE.json.

    The completeness gate checks report coverage against this
    authoritative set (see audit-quickref Step 1), so the agent cannot
    pass by shrinking its own denominator. With --platform/--country it
    prints a ready combo object; otherwise just the id array.
    """
    _wb, _ws, _header, data = load(args.file)
    active_ids = [
        str(r[Col.CAMPAIGN_ID]).strip()
        for r in data
        if is_entity(r, 'campaign')
        and state_api(r[Col.STATE]) == STATE_ENABLED
        and r[Col.CAMPAIGN_ID] not in (None, '')
    ]
    active_ids = list(dict.fromkeys(active_ids))  # de-dupe, preserve order
    if args.platform and args.country:
        out = {
            'platform': args.platform,
            'country': args.country,
            'active_ids': active_ids,
        }
    else:
        out = active_ids
    print(json.dumps(out, ensure_ascii=False, indent=2))


def main():
    ap = argparse.ArgumentParser(description=__doc__.split('\n')[0])
    sub = ap.add_subparsers(dest='cmd', required=True)

    p = sub.add_parser('inspect', help='summarise a bulk export')
    p.add_argument('file')
    p.set_defaults(func=cmd_inspect)

    p = sub.add_parser(
        'scope', help='print active Campaign ids for AUDIT_SCOPE.json'
    )
    p.add_argument('file')
    p.add_argument('--platform', help='e.g. amazon (emit a full combo obj)')
    p.add_argument('--country', help='e.g. SA (emit a full combo obj)')
    p.set_defaults(func=cmd_scope)

    p = sub.add_parser(
        'clone-campaign', help='emit Create rows for a new manual campaign'
    )
    p.add_argument('file')
    p.add_argument(
        '--src',
        required=True,
        help='exact source campaign name (keyword source)',
    )
    p.add_argument('--new', required=True, help='new campaign name')
    p.add_argument('--ad-group', help='ad group name (default: new name)')
    p.add_argument('--sku', required=True, help='real seller SKU')
    p.add_argument('--asin', help='ASIN (informational only)')
    p.add_argument('--daily-budget', type=float, default=1.0)
    p.add_argument('--default-bid', type=float, default=0.75)
    p.add_argument(
        '--keyword-bid', type=float, help='override bid for all cloned keywords'
    )
    p.add_argument('--bidding-strategy', default='Dynamic bids - down only')
    p.add_argument(
        '--start-date',
        help='campaign start date YYYYMMDD (required by Amazon; '
        'default: today)',
    )
    p.add_argument('--out')
    p.set_defaults(func=cmd_clone_campaign)

    p = sub.add_parser('bid-update', help='emit Update rows for bids')
    p.add_argument('file')
    p.add_argument('--campaign', help='limit to one campaign name')
    p.add_argument('--scale', type=float, help='multiply each bid by F')
    p.add_argument('--set-bid', type=float, help='set every bid to V')
    p.add_argument('--out')
    p.set_defaults(func=cmd_bid_update)

    p = sub.add_parser(
        'archive-campaign', help='emit an Archive row to disable a campaign'
    )
    p.add_argument('file')
    p.add_argument('--campaign', required=True, help='exact campaign name')
    p.add_argument('--out')
    p.set_defaults(func=cmd_archive_campaign)

    args = ap.parse_args()
    args.func(args)


if __name__ == '__main__':
    main()
