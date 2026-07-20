"""Unit tests for the amazon-listing skill's scripts/listing_bulk.py.

Pins the invariants that make the flat-file listing tool trustworthy:

* **Locale-generality** — the tool keys every field by its field API
  name (the row containing `item_sku`), never by the localised label
  row above it. The synthetic template here puts **Chinese** labels in
  that row and asserts every operation still works.
* **The operation column** — `create` leaves `update_delete` blank,
  and `update` / `partialupdate` / `delete` write their tokens.
* **Parent-child shape** — a Parent row carries the variation theme and
  no offer; a Child row carries `parent_sku`, the differentiator, and
  the price; the Parent is not warned about child-level required fields.
* **Validation** — an out-of-enum value and a missing required field
  each raise a warning (never a hard failure), so an unseen-but-valid
  token from a new category still writes.

All test data is fabricated — no real store, brand, SKU, or ASIN.
"""

import importlib.util
import json
from pathlib import Path
import sys

import pytest

from app.ai.skill_review import parse_skill_review

pytestmark = pytest.mark.unit

openpyxl = pytest.importorskip('openpyxl')
from openpyxl.comments import Comment  # noqa: E402  (after importorskip guard)

_SKILL_PATH = (
    Path(__file__).resolve().parents[2]
    / 'app'
    # Live skill tree (config.SKILLS_SUBDIR); matches the ads-bulk test.
    # ``app/skills`` is the frozen 0.12.x copy and is not what ships.
    / 'skills_v2'
    / 'amazon-listing'
    / 'scripts'
    / 'listing_bulk.py'
)
_spec = importlib.util.spec_from_file_location('listing_bulk', _SKILL_PATH)
listing_bulk = importlib.util.module_from_spec(_spec)
sys.modules['listing_bulk'] = listing_bulk
_spec.loader.exec_module(listing_bulk)


# Field API names in column order. Localised labels sit one row above
# them (Chinese here) — the tool must ignore the labels entirely.
FIELDS = [
    'feed_product_type',
    'item_sku',
    'update_delete',
    'brand_name',
    'external_product_id',
    'external_product_id_type',
    'item_name',
    'title_differentiation',
    'product_description',
    'recommended_browse_nodes',
    'main_image_url',
    'parent_child',
    'relationship_type',
    'variation_theme',
    'parent_sku',
    'color_name',
    'batteries_required',
    'are_batteries_included',
    'fulfillment_availability#1.quantity',
    'purchasable_offer[marketplace_id=MKTSA]#1.our_price#1.schedule#1.value_with_tax',
]
ZH_LABELS = {
    'feed_product_type': '产品类型',
    'item_sku': '卖家SKU',
    'update_delete': '更新删除',
    'brand_name': '品牌名称',
    'item_name': '商品名称',
    'color_name': '颜色',
    'parent_child': '父子关系',
    'variation_theme': '变体主题',
}
REQUIRED = {
    'feed_product_type',
    'item_sku',
    'brand_name',
    'item_name',
    'product_description',
    'recommended_browse_nodes',
    'external_product_id',
    'external_product_id_type',
    'main_image_url',
    'batteries_required',
    'battery_type',  # battery_type: conditional
    'fulfillment_availability#1.quantity',
    'purchasable_offer[marketplace_id=MKTSA]#1.our_price#1.schedule#1.value_with_tax',
}
ENUMS = {
    'update_delete': ['Update', 'partialupdate', 'delete'],
    'parent_child': ['Parent', 'Child'],
    'relationship_type': ['Variation'],
    'variation_theme': ['Color', 'Size', 'color-size'],
    'external_product_id_type': ['asin', 'upc', 'ean', 'gtin'],
    'recommended_browse_nodes': ['111', '222', '333'],
}


def _make_template(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = listing_bulk.TEMPLATE_SHEET
    # Row 1: signature/group row (decorative). Row 2: localised labels.
    # Row 3: field API names. Row 4+: data (empty in a blank template).
    ws.append(['TemplateType=fptcustom', 'Version=test'])
    ws.append([ZH_LABELS.get(f, f) for f in FIELDS])
    ws.append(list(FIELDS))

    # Data Definitions: header at row 2, then per-field rows.
    dd = wb.create_sheet(listing_bulk.DEFN_SHEET)
    dd.append(['How to complete your inventory template'])
    dd.append([
        'Group Name',
        'Field Name',
        'Local Label Name',
        'Definition and Use',
        'Accepted Values',
        'Example',
        'Required?',
    ])
    for f in FIELDS + ['battery_type']:
        dd.append([
            '',
            f,
            ZH_LABELS.get(f, f),
            '',
            '',
            '',
            'Required' if f in REQUIRED else 'Optional',
        ])

    # Dropdown Lists: field names on row 3, tokens below each column.
    dl = wb.create_sheet(listing_bulk.DROPDOWN_SHEET)
    dl.append([])
    dl.append([])
    enum_fields = list(ENUMS)
    dl.append(enum_fields)
    for i in range(max(len(v) for v in ENUMS.values())):
        dl.append([
            ENUMS[f][i] if i < len(ENUMS[f]) else None for f in enum_fields
        ])
    wb.save(path)


@pytest.fixture
def template(tmp_path):
    p = tmp_path / 'template.xlsx'
    _make_template(str(p))
    return str(p)


def _run(args):
    ns = listing_bulk.argparse.Namespace
    parser_map = {
        'inspect': listing_bulk.cmd_inspect,
        'fill': listing_bulk.cmd_fill,
        'parse-feedback': listing_bulk.cmd_parse_feedback,
    }
    # Route through the real argv parser so the test exercises main().
    old = sys.argv
    sys.argv = ['listing_bulk.py'] + args
    try:
        listing_bulk.main()
    finally:
        sys.argv = old
    return ns, parser_map  # unused; kept for clarity


def _read_rows(path):
    wb = openpyxl.load_workbook(path)
    ws = wb[listing_bulk.TEMPLATE_SHEET]
    names = [c.value for c in ws[3]]
    idx = {n: i for i, n in enumerate(names)}
    rows = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        if any(c is not None for c in r):
            rows.append({n: r[idx[n]] for n in names})
    return rows


def _spec(tmp_path, rows):
    p = tmp_path / 'spec.json'
    p.write_text(json.dumps(rows), encoding='utf-8')
    return str(p)


def test_header_found_by_field_name_not_localised_label(template):
    wb = openpyxl.load_workbook(template)
    ws = wb[listing_bulk.TEMPLATE_SHEET]
    # Header row is row 3 (field names), NOT row 2 (Chinese labels).
    assert listing_bulk._find_header_row(ws) == 3
    required = listing_bulk._load_required_fields(wb)
    assert 'item_sku' in required and 'brand_name' in required
    valid = listing_bulk._load_valid_values(wb)
    assert valid['parent_child'] == {'parent', 'child'}
    assert 'delete' in valid['update_delete']


def test_fill_operation_column(template, tmp_path):
    spec = {
        'product_type': 'socks',
        'brand': 'ACME',
        'rows': [
            {
                'sku': 'W-1',
                'operation': 'create',
                'fields': {'item_name': 'ACME Thing'},
            },
            {
                'sku': 'W-2',
                'operation': 'update',
                'fields': {'item_name': 'ACME Thing 2'},
            },
            {
                'sku': 'W-3',
                'operation': 'partialupdate',
                'fields': {'item_name': 'ACME Thing 3'},
            },
            {'sku': 'W-4', 'operation': 'delete'},
        ],
    }
    out = str(tmp_path / 'out.xlsx')
    _run(['fill', template, '--spec', _spec(tmp_path, spec), '--out', out])
    rows = _read_rows(out)
    by_sku = {r['item_sku']: r for r in rows}
    assert by_sku['W-1']['update_delete'] in (None, '')  # create = blank
    assert by_sku['W-2']['update_delete'] == 'Update'
    assert by_sku['W-3']['update_delete'] == 'partialupdate'
    assert by_sku['W-4']['update_delete'] == 'delete'
    # brand default applied to every row.
    assert all(r['brand_name'] == 'ACME' for r in rows)


def test_fill_drops_item_highlight_when_title_too_long(
    template, tmp_path, capsys
):
    # Item Highlight (title_differentiation) is only valid when item_name
    # is <= 75 chars; a longer title makes Amazon reject it with 100476
    # (SUCCESS OTHER). fill drops the optional highlight so the SKU lands
    # clean, and warns. Regression for the long-title over-claim.
    long_title = 'A' * 90
    spec = {
        'product_type': 'socks',
        'brand': 'ACME',
        'rows': [
            {
                'sku': 'W-LONG',
                'operation': 'create',
                'fields': {
                    'item_name': long_title,
                    'title_differentiation': 'Pure White',
                },
            }
        ],
    }
    out = str(tmp_path / 'out.xlsx')
    _run(['fill', template, '--spec', _spec(tmp_path, spec), '--out', out])
    row = {r['item_sku']: r for r in _read_rows(out)}['W-LONG']
    assert row['title_differentiation'] in (None, '')
    err = capsys.readouterr().err
    assert 'Item Highlight' in err and '100476' in err


def test_fill_keeps_item_highlight_when_title_fits(template, tmp_path):
    # A short title is under the limit, so a genuine Item Highlight is
    # written unchanged -- the guard only strips the poison case.
    spec = {
        'product_type': 'socks',
        'brand': 'ACME',
        'rows': [
            {
                'sku': 'W-SHORT',
                'operation': 'create',
                'fields': {
                    'item_name': 'ACME Socks',
                    'title_differentiation': 'Breathable',
                },
            }
        ],
    }
    out = str(tmp_path / 'out.xlsx')
    _run(['fill', template, '--spec', _spec(tmp_path, spec), '--out', out])
    row = {r['item_sku']: r for r in _read_rows(out)}['W-SHORT']
    assert row['title_differentiation'] == 'Breathable'


def test_parent_child_structure(template, tmp_path):
    spec = {
        'product_type': 'socks',
        'brand': 'ACME',
        'rows': [
            {
                'sku': 'P-1',
                'operation': 'create',
                'parentage': 'Parent',
                'variation_theme': 'Color',
                'fields': {
                    'relationship_type': 'Variation',
                    'item_name': 'ACME Socks',
                    'product_description': 'desc',
                    'recommended_browse_nodes': '111',
                    'batteries_required': 'No',
                    'are_batteries_included': 'No',
                },
            },
            {
                'sku': 'P-1-WHT',
                'operation': 'create',
                'parentage': 'Child',
                'parent_sku': 'P-1',
                'variation_theme': 'Color',
                'fields': {
                    'relationship_type': 'Variation',
                    'color_name': 'White',
                    'item_name': 'ACME Socks White',
                    'product_description': 'desc',
                    'recommended_browse_nodes': '111',
                    'external_product_id': '00012345678905',
                    'external_product_id_type': 'upc',
                    'main_image_url': 'https://example.com/w.jpg',
                    'batteries_required': 'No',
                    'are_batteries_included': 'No',
                    'fulfillment_availability#1.quantity': '100',
                    'purchasable_offer[marketplace_id=MKTSA]#1.our_price'
                    '#1.schedule#1.value_with_tax': '29.00',
                },
            },
        ],
    }
    out = str(tmp_path / 'out.xlsx')
    _run(['fill', template, '--spec', _spec(tmp_path, spec), '--out', out])
    rows = _read_rows(out)
    parent = next(r for r in rows if r['item_sku'] == 'P-1')
    child = next(r for r in rows if r['item_sku'] == 'P-1-WHT')
    assert parent['parent_child'] == 'Parent'
    assert parent['variation_theme'] == 'Color'
    assert parent['parent_sku'] in (None, '')  # parent has no parent
    price_col = (
        'purchasable_offer[marketplace_id=MKTSA]#1.our_price'
        '#1.schedule#1.value_with_tax'
    )
    assert parent[price_col] in (None, '')  # offer is child-level
    assert child['parent_child'] == 'Child'
    assert child['parent_sku'] == 'P-1'
    assert child['color_name'] == 'White'
    assert str(child[price_col]) == '29.00'


def test_parent_not_warned_for_child_level_required(template, tmp_path, capsys):
    """A Parent row omitting offer/price/image/product-id must NOT be
    flagged — those are child-level. A Child omitting them must be."""
    spec = {
        'product_type': 'socks',
        'brand': 'ACME',
        'rows': [
            {
                'sku': 'P-1',
                'operation': 'create',
                'parentage': 'Parent',
                'variation_theme': 'Color',
                'fields': {
                    'relationship_type': 'Variation',
                    'item_name': 'ACME Socks',
                    'product_description': 'desc',
                    'recommended_browse_nodes': '111',
                    'batteries_required': 'No',
                    'are_batteries_included': 'No',
                },
            },
        ],
    }
    out = str(tmp_path / 'out.xlsx')
    _run(['fill', template, '--spec', _spec(tmp_path, spec), '--out', out])
    err = capsys.readouterr().err
    # Parent: no complaint about child-level fields...
    assert 'external_product_id' not in err
    assert 'our_price' not in err
    assert 'main_image_url' not in err
    # ...and no complaint about battery_type (declared no batteries).
    assert 'battery_type' not in err


def test_out_of_enum_and_missing_required_warn_not_fail(
    template, tmp_path, capsys
):
    spec = {
        'product_type': 'socks',
        'brand': 'ACME',
        'rows': [
            {
                'sku': 'W-1',
                'operation': 'create',
                'variation_theme': 'PurpleHaze',  # not in enum
                'fields': {'relationship_type': 'Variation'},
            },  # missing lots
        ],
    }
    out = str(tmp_path / 'out.xlsx')
    _run(['fill', template, '--spec', _spec(tmp_path, spec), '--out', out])
    err = capsys.readouterr().err
    assert 'PurpleHaze' in err  # enum warning
    assert 'not in template valid values' in err
    assert 'missing required field' in err  # required warning
    # File still written despite warnings.
    assert Path(out).exists()
    assert _read_rows(out)[0]['variation_theme'] == 'PurpleHaze'


def test_asin_folds_into_product_id(template, tmp_path):
    spec = {
        'product_type': 'socks',
        'brand': 'ACME',
        'rows': [
            {
                'sku': 'W-1',
                'operation': 'update',
                'asin': 'B0ABCDE123',
                'fields': {'item_name': 'x'},
            },
        ],
    }
    out = str(tmp_path / 'out.xlsx')
    _run(['fill', template, '--spec', _spec(tmp_path, spec), '--out', out])
    row = _read_rows(out)[0]
    assert row['external_product_id'] == 'B0ABCDE123'
    assert row['external_product_id_type'] == 'asin'


def test_fill_clears_preexisting_data_rows(template, tmp_path):
    """A reused template with a stale data row must not upload it —
    fill clears the data area before writing the spec rows."""
    wb = openpyxl.load_workbook(template)
    ws = wb[listing_bulk.TEMPLATE_SHEET]
    # Row 4 (first data row): a stale SKU from a prior run.
    ws.append(['socks', 'STALE-SKU'] + [''] * (len(FIELDS) - 2))
    wb.save(template)

    spec = {
        'product_type': 'socks',
        'brand': 'ACME',
        'rows': [
            {
                'sku': 'W-NEW',
                'operation': 'create',
                'fields': {'item_name': 'x'},
            },
        ],
    }
    out = str(tmp_path / 'out.xlsx')
    _run(['fill', template, '--spec', _spec(tmp_path, spec), '--out', out])
    skus = [r['item_sku'] for r in _read_rows(out)]
    assert 'STALE-SKU' not in skus
    assert skus == ['W-NEW']


def test_parse_feedback(tmp_path, capsys):
    report = tmp_path / 'report.txt'
    report.write_text(
        'original-record-number\tsku\terror-type\terror-code\t'
        'error-message\n'
        '4\tW-1\tError\t8541\tMissing required field external_product_id\n'
        '5\tW-2\tWarning\t90220\tImage will be reviewed\n',
        encoding='utf-8',
    )
    with pytest.raises(SystemExit) as ei:  # exit 1 because an error
        _run(['parse-feedback', str(report)])
    assert ei.value.code == 1
    out = capsys.readouterr().out
    assert 'W-1' in out and 'external_product_id' in out
    assert '1 error(s), 1 warning(s)' in out


def test_fill_emits_tsv_upload_file(template, tmp_path):
    """fill writes a tab-delimited .txt sibling (the upload artefact) with
    a uniform column count. Uploading the .xlsm triggers Amazon's 90502
    FATAL; the .txt is what reaches content validation."""
    spec = {
        'product_type': 'socks',
        'brand': 'ACME',
        'rows': [
            {
                'sku': 'W-1',
                'operation': 'create',
                'fields': {'item_name': 'ACME Thing'},
            },
            {
                'sku': 'W-2',
                'operation': 'create',
                'fields': {'item_name': 'ACME Thing 2'},
            },
        ],
    }
    out = str(tmp_path / 'out.xlsx')
    _run(['fill', template, '--spec', _spec(tmp_path, spec), '--out', out])
    txt = tmp_path / 'out.txt'
    assert txt.exists()
    lines = [ln for ln in txt.read_text(encoding='utf-8').split('\n') if ln]
    grid = [ln.split('\t') for ln in lines]
    # Uniform width -> no ragged rows -> no column drift on Amazon's side.
    assert len({len(r) for r in grid}) == 1
    names = grid[2]  # row 3 = field API names
    assert 'item_sku' in names
    sku_col = names.index('item_sku')
    assert {'W-1', 'W-2'} <= {r[sku_col] for r in grid[3:]}


def test_enum_value_canonicalised_to_template_case(template, tmp_path):
    """A spec enum value in the wrong case is written in the template's
    exact case. Amazon is case-strict on some fields (verified live:
    apparel_size_system accepts UAE/KSA but rejects uae/ksa)."""
    spec = {
        'product_type': 'socks',
        'brand': 'ACME',
        'rows': [
            {
                'sku': 'W-1',
                'operation': 'create',
                'variation_theme': 'color',  # lowercase; template has 'Color'
                'fields': {'item_name': 'ACME Thing'},
            },
        ],
    }
    out = str(tmp_path / 'out.xlsx')
    _run(['fill', template, '--spec', _spec(tmp_path, spec), '--out', out])
    rows = _read_rows(out)
    assert rows[0]['variation_theme'] == 'Color'


def test_parse_feedback_extracts_cell_comments(tmp_path):
    """parse-feedback's extractor surfaces the per-cell 批注 (the
    field-level verdict) from the report's Template tab, split into
    individual ERROR/WARNING lines with Excel's _x000d_ marker cleaned.
    This is the engine of the self-correct loop."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = listing_bulk.TEMPLATE_SHEET
    ws.append(['signature'])  # row 1
    ws.append(['Seller SKU', 'Colour'])  # row 2 localised labels
    ws.append(['item_sku', 'color_name'])  # row 3 field API names
    ws.append(['W-1', 'White'])  # row 4 data
    # Amazon stacks several messages in one comment, split by _x000d_.
    ws.cell(row=4, column=1).comment = Comment(
        'ERROR : missing standard_product_id_x000d_WARNING : missing material',
        'Amazon',
    )
    p = tmp_path / 'report.xlsx'
    wb.save(str(p))

    errs = list(listing_bulk._template_cell_errors(str(p)))
    msgs = [m for _, _, m in errs]
    assert all(sku == 'W-1' for sku, _, _ in errs)
    assert any(
        m.startswith('ERROR') and 'standard_product_id' in m for m in msgs
    )
    assert any(m.startswith('WARNING') and 'material' in m for m in msgs)
    # _x000d_ artifact must be cleaned (split into separate lines).
    assert not any('_x000d_' in m for m in msgs)


# --- offer routing to the target marketplace (multi-marketplace bug) ---

_SA = 'A17E79C6D8DWNP'
_AE = 'A2VIGQ35RCS4UG'
_SA_PRICE = f'purchasable_offer[marketplace_id={_SA}]#1.our_price#1.schedule#1.value_with_tax'
_AE_PRICE = f'purchasable_offer[marketplace_id={_AE}]#1.our_price#1.schedule#1.value_with_tax'


def _make_mkt_template(path):
    """A template with a SEPARATE offer-price column per marketplace."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = listing_bulk.TEMPLATE_SHEET
    fields = [
        'feed_product_type',
        'item_sku',
        'brand_name',
        'update_delete',
        'item_name',
        'parent_child',
        'relationship_type',
        'variation_theme',
        'parent_sku',
        'color_name',
        'fulfillment_availability#1.quantity',
        _SA_PRICE,
        _AE_PRICE,
    ]
    ws.append(['TemplateType=fptcustom'])
    ws.append(['label:' + f for f in fields])  # localised label row
    ws.append(list(fields))  # field API-name row (header)
    dd = wb.create_sheet(listing_bulk.DEFN_SHEET)
    dd.append(['x'])
    dd.append([
        'Group Name',
        'Field Name',
        'Local Label Name',
        'Definition and Use',
        'Accepted Values',
        'Example',
        'Required?',
    ])
    for f in fields:
        # A multi-marketplace template marks a *different* marketplace's
        # offer block Required than the one we're listing on -- exactly the
        # trap that silently fills the wrong marketplace's price.
        dd.append([
            '',
            f,
            f,
            '',
            '',
            '',
            'Required' if f == _AE_PRICE else 'Optional',
        ])
    wb.save(path)


@pytest.fixture
def mkt_template(tmp_path):
    p = tmp_path / 'mkt.xlsx'
    _make_mkt_template(str(p))
    return str(p)


def test_fill_routes_bare_our_price_to_target_marketplace(
    mkt_template, tmp_path
):
    spec = _spec(
        tmp_path,
        {
            'marketplace': 'SA',
            'product_type': 'socks',
            'rows': [
                {
                    'sku': 'K-WHT',
                    'parentage': 'Child',
                    'variation_theme': 'Color',
                    'fields': {
                        'item_name': 'x',
                        'color_name': 'White',
                        'feed_product_type': 'socks',
                        'our_price': '19.99',
                    },
                }
            ],
        },
    )
    out = str(tmp_path / 'out.xlsx')
    _run(['fill', mkt_template, '--spec', spec, '--out', out])
    row = _read_rows(out)[0]
    assert row[_SA_PRICE] == '19.99'  # target marketplace got the price
    assert row[_AE_PRICE] in (None, '')  # the wrong block stayed empty


def test_fill_routes_row_level_our_price_and_quantity(mkt_template, tmp_path):
    # The skill says put "a bare our_price / quantity on each child" -- an
    # agent naturally puts them at the ROW level (sibling to `fields`), not
    # inside `fields`. Both placements must route, else the offer column
    # comes out empty and the agent thrashes hand-picking columns.
    spec = _spec(
        tmp_path,
        {
            'marketplace': 'SA',
            'product_type': 'socks',
            'rows': [
                {
                    'sku': 'K-WHT',
                    'parentage': 'Child',
                    'variation_theme': 'Color',
                    'our_price': '19.99',  # row-level, not in fields
                    'quantity': '100',  # row-level, not in fields
                    'fields': {
                        'item_name': 'x',
                        'color_name': 'White',
                        'feed_product_type': 'socks',
                    },
                }
            ],
        },
    )
    out = str(tmp_path / 'out.xlsx')
    _run(['fill', mkt_template, '--spec', spec, '--out', out])
    row = _read_rows(out)[0]
    assert row[_SA_PRICE] == '19.99'
    assert row[_AE_PRICE] in (None, '')


def test_fill_warns_when_target_marketplace_has_no_offer(
    mkt_template, tmp_path, capsys
):
    # Agent hand-picked the AE column but is listing on SA. We DON'T silently
    # move it (that confused agents) -- leave it as written and WARN that SA
    # has no offer (=> Missing offer).
    spec = _spec(
        tmp_path,
        {
            'marketplace': 'SA',
            'product_type': 'socks',
            'rows': [
                {
                    'sku': 'K-WHT',
                    'parentage': 'Child',
                    'variation_theme': 'Color',
                    'fields': {
                        'item_name': 'x',
                        'color_name': 'White',
                        'feed_product_type': 'socks',
                        _AE_PRICE: '19.99',
                    },
                }
            ],
        },
    )
    out = str(tmp_path / 'out.xlsx')
    _run(['fill', mkt_template, '--spec', spec, '--out', out])
    row = _read_rows(out)[0]
    assert row[_AE_PRICE] == '19.99'  # explicit column left as written
    assert row[_SA_PRICE] in (None, '')
    assert 'Missing offer' in capsys.readouterr().err


def test_fill_errors_when_price_set_without_marketplace(mkt_template, tmp_path):
    spec = _spec(
        tmp_path,
        {
            'product_type': 'socks',
            'rows': [
                {
                    'sku': 'K-WHT',
                    'parentage': 'Child',
                    'fields': {'item_name': 'x', 'our_price': '19.99'},
                }
            ],
        },
    )
    out = str(tmp_path / 'out.xlsx')
    with pytest.raises(SystemExit) as e:
        _run(['fill', mkt_template, '--spec', spec, '--out', out])
    assert 'marketplace' in str(e.value)


def test_fill_auto_derives_relationship_type_for_variation(
    mkt_template, tmp_path
):
    # A child with parent_sku + variation_theme but NO relationship_type must
    # still get relationship_type=Variation (else Amazon errors
    # "relationship_type = null" and the child never creates).
    spec = _spec(
        tmp_path,
        {
            'marketplace': 'SA',
            'product_type': 'socks',
            'rows': [
                {
                    'sku': 'K-P',
                    'parentage': 'Parent',
                    'variation_theme': 'Color',
                    'fields': {'item_name': 'p'},
                },
                {
                    'sku': 'K-WHT',
                    'parentage': 'Child',
                    'parent_sku': 'K-P',
                    'variation_theme': 'Color',
                    'fields': {
                        'item_name': 'w',
                        'color_name': 'White',
                        'our_price': '9.99',
                    },
                },
            ],
        },
    )
    out = str(tmp_path / 'out.xlsx')
    _run(['fill', mkt_template, '--spec', spec, '--out', out])
    rows = {r['item_sku']: r for r in _read_rows(out)}
    assert rows['K-P']['relationship_type'] == 'Variation'
    assert rows['K-WHT']['relationship_type'] == 'Variation'


def test_marketplace_table_is_global():
    # The country->id table must cover Amazon's global marketplaces, not
    # just the few we happened to list on. Spot-check one per region and a
    # healthy total so a truncated table fails loudly.
    ids = listing_bulk.MARKETPLACE_IDS
    for code, mid in {
        'US': 'ATVPDKIKX0DER',
        'UK': 'A1F83G8C2ARO7P',
        'JP': 'A1VC38T7YXB528',
        'AU': 'A39IBJ37TRP1C6',
        'SA': 'A17E79C6D8DWNP',
        'AE': 'A2VIGQ35RCS4UG',
        'BR': 'A2Q3Y263D00KWC',
    }.items():
        assert ids[code] == mid
    assert len(ids) >= 20  # NA + EU + ME + APAC, not a partial subset


def test_resolve_marketplace_by_code_and_raw_id():
    r = listing_bulk._resolve_marketplace_id
    assert r('SA') == _SA
    assert r('sa') == _SA  # case-insensitive country code
    assert r(_AE) == _AE  # a raw id passes through unchanged
    assert r(None) is None  # nothing supplied, no template hint


def test_marketplace_ids_read_from_template_columns():
    cols = {'item_sku': 1, _SA_PRICE: 2, _AE_PRICE: 3}
    assert listing_bulk._marketplace_ids_in_template(cols) == [_SA, _AE]
    assert listing_bulk._marketplace_ids_in_template({'item_sku': 1}) == []


def test_resolve_auto_detects_single_marketplace_template():
    r = listing_bulk._resolve_marketplace_id
    # No marketplace given, but the template offers exactly one -> use it.
    assert r(None, [_SA]) == _SA
    # A marketplace code Amazon added after our table, disambiguated by a
    # single-marketplace template -> still resolves (stays general).
    assert r('ZA', [_SA]) == _SA
    # Ambiguous: unknown code + multi-marketplace template -> fail loudly.
    with pytest.raises(SystemExit):
        r('ZA', [_SA, _AE])


# A real multi-marketplace template interleaves each marketplace's
# fulfillment_availability#N group with its offer block: AE (#1) then SA
# (#2). Column order matters -- the resolver picks the group by position.
_AE_QTY = 'fulfillment_availability#1.quantity'
_SA_QTY = 'fulfillment_availability#2.quantity'
_MKT_COLS = {
    'item_sku': [2],
    _AE_QTY: [165],
    _AE_PRICE: [171],
    _SA_QTY: [181],
    _SA_PRICE: [186],
}


def test_fulfillment_index_maps_to_the_marketplaces_own_group():
    f = listing_bulk._fulfillment_index_for_marketplace
    assert f(_MKT_COLS, _AE) == 1  # AE's stock group precedes AE's offer
    assert f(_MKT_COLS, _SA) == 2  # SA's stock group precedes SA's offer


def test_route_sends_quantity_to_target_marketplace_group():
    # Listing on SA: a bare `quantity` must land in SA's group (#2), NOT the
    # first group (#1 = AE) -- else the SA offer has no stock and never lives.
    fields = {'our_price': '19.99', 'quantity': '100'}
    listing_bulk._route_offer_price(fields, _MKT_COLS, _SA, 0, 'K-WHT', [])
    assert fields[_SA_QTY] == '100'
    assert _AE_QTY not in fields
    assert fields[_SA_PRICE] == '19.99'


def test_route_remaps_wrong_index_fulfillment_to_target():
    # The agent wrote fulfillment_availability#1 (AE) but is listing on SA;
    # normalise it to SA's group (#2).
    fields = {_AE_QTY: '100', 'our_price': '19.99'}
    listing_bulk._route_offer_price(fields, _MKT_COLS, _SA, 0, 'K-WHT', [])
    assert fields[_SA_QTY] == '100'
    assert _AE_QTY not in fields


def test_route_sends_channel_code_to_target_group():
    # A bare fulfillment_channel_code routes to the target marketplace's
    # fulfillment group (SA = #2), paired with its quantity -- a group needs
    # both or Amazon rejects it ("does not have enough values").
    fields = {
        'our_price': '19.99',
        'quantity': '100',
        'fulfillment_channel_code': 'DEFAULT',
    }
    warnings = []
    listing_bulk._route_offer_price(
        fields, _MKT_COLS, _SA, 0, 'K-WHT', warnings
    )
    assert fields[_SA_QTY] == '100'
    assert (
        fields['fulfillment_availability#2.fulfillment_channel_code']
        == 'DEFAULT'
    )
    assert warnings == []  # both quantity and code present -> no warning


def test_route_warns_when_group_has_quantity_but_no_channel_code():
    # quantity routed to the target group but no channel code -> the exact
    # cross-marketplace rejection; warn so the caller supplies the code.
    fields = {'our_price': '19.99', 'quantity': '100'}
    warnings = []
    listing_bulk._route_offer_price(
        fields, _MKT_COLS, _SA, 0, 'K-WHT', warnings
    )
    assert any('fulfillment_channel_code' in w for w in warnings)


# --- unified (NGS "Beta Product Spreadsheet") template dialect ----------
#
# The current Seller Central template. Its field API names are decorated
# (`contribution_sku#1.value`, `::record_action`, marketplace-scoped
# parentage/parent-link/offer with an `[audience=ALL]` insert), its
# field-name row sits at Excel row 5 (not 3), and it ships PREFILLED rows
# (an Example SKU + a "do not delete this row" instruction). The legacy
# tool keyed on `item_sku` and crashed here, which forced agents to
# hand-roll the file -- and a hand-rolled file uploaded the prefilled
# Example/instruction rows as real SKUs (verified: a live run got "1/8"
# because those junk rows shipped). These tests pin that the SAME friendly
# spec now drives this dialect and clears the prefilled rows.

_UNI_SKU = 'contribution_sku#1.value'
_UNI_PT = 'product_type#1.value'
_UNI_OP = '::record_action'
_UNI_PARENTAGE = f'parentage_level[marketplace_id={_SA}]#1.value'
_UNI_PARENT_SKU = (
    f'child_parent_sku_relationship[marketplace_id={_SA}]#1.parent_sku'
)
_UNI_THEME = 'variation_theme#1.name'
_UNI_NAME = f'item_name[marketplace_id={_SA}][language_tag=en_AE]#1.value'
_UNI_BRAND = f'brand[marketplace_id={_SA}][language_tag=en_AE]#1.value'
_UNI_ID_TYPE = 'amzn1.volt.ca.product_id_type'
_UNI_ID_VALUE = 'amzn1.volt.ca.product_id_value'
_UNI_COLOR = f'color[marketplace_id={_SA}][language_tag=en_AE]#1.value'
_UNI_QTY = 'fulfillment_availability#1.quantity'
# The offer column carries the `[audience=ALL]` insert a fixed template
# string can't match -- the routing must find it structurally.
_UNI_PRICE = (
    f'purchasable_offer[marketplace_id={_SA}][audience=ALL]'
    '#1.our_price#1.schedule#1.value_with_tax'
)

# Column order matters: stock group must precede the offer block.
_UNI_FIELDS = [
    _UNI_SKU,
    _UNI_PT,
    _UNI_OP,
    _UNI_PARENTAGE,
    _UNI_PARENT_SKU,
    _UNI_THEME,
    _UNI_NAME,
    _UNI_BRAND,
    _UNI_ID_TYPE,
    _UNI_ID_VALUE,
    _UNI_COLOR,
    _UNI_QTY,
    _UNI_PRICE,
]
_UNI_REQUIRED = {
    _UNI_SKU,
    _UNI_PT,
    _UNI_NAME,
    _UNI_BRAND,
    _UNI_PRICE,
    _UNI_COLOR,
}
_UNI_ENUMS = {
    _UNI_OP: [
        'Create or Replace (Full Update)',
        'Edit (Partial Update)',
        'Delete',
    ],
    _UNI_PARENTAGE: ['Parent', 'Child'],
    _UNI_THEME: ['COLLECTION_ITEM', 'COLOR', 'SIZE'],
    _UNI_ID_TYPE: ['EAN', 'GTIN', 'UPC', 'ASIN', 'GTIN Exempt'],
}


def _make_unified_template(path):
    """A synthetic unified template mirroring the real geometry: a 5-row
    header (settings / instructions / group / localised labels / field API
    names) and PREFILLED Example + instruction data rows that fill MUST
    clear."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = listing_bulk.TEMPLATE_SHEET
    # row 1 settings — like the real template, carries the geometry
    # (labelRow/attributeRow/dataRow). dataRow=8 means rows 6-7 are a
    # skipped example region and data must start at row 8.
    ws.append([
        'settings=feedType=256&labelRow=4&attributeRow=5&dataRow=8'
        '&contributorId=amzn1'
    ])
    ws.append(['Use ENGLISH. DO NOT modify or delete the colored head'])
    ws.append(['Listing Identity'])  # row 3 group names
    ws.append(['SKU', 'Product Type', 'Listing Action'])  # row 4 labels (zh)
    ws.append(list(_UNI_FIELDS))  # row 5 field API names <-- keyed
    # Row 6+: Amazon's prefilled Example row + a "do not delete" row.
    example = {
        _UNI_SKU: 'EXAMPLE-SKU',
        _UNI_OP: '(Default) Create or Replace',
        _UNI_PARENTAGE: 'Parent',
        _UNI_PARENT_SKU: 'EXAMPLE-SKU',
        _UNI_THEME: 'SIZE',
        _UNI_ID_TYPE: 'UPC',
        _UNI_ID_VALUE: '100000000001',
        _UNI_PRICE: '9.00',
    }
    idx = {f: i for i, f in enumerate(_UNI_FIELDS)}
    row6 = [''] * len(_UNI_FIELDS)
    for f, v in example.items():
        row6[idx[f]] = v
    ws.append(row6)
    row7 = [''] * len(_UNI_FIELDS)
    row7[0] = "We've prefilled attributes. Please do not delete this row."
    ws.append(row7)

    # Unified Data Definitions has NO 'Definition and Use' column, so
    # 'Required?' sits at col 6 (index 5), not col 7 — the tool must find
    # it by header name, not a fixed index.
    dd = wb.create_sheet(listing_bulk.DEFN_SHEET)
    dd.append(['How to complete your inventory template'])
    dd.append([
        'Group Name',
        'Field Name',
        'Local Label Name',
        'Accepted Values',
        'Example',
        'Required?',
    ])
    for f in _UNI_FIELDS:
        dd.append(['', f, f, '', '', 'Required' if f in _UNI_REQUIRED else ''])

    dl = wb.create_sheet(listing_bulk.DROPDOWN_SHEET)
    dl.append([])
    dl.append([])
    enum_fields = list(_UNI_ENUMS)
    dl.append(enum_fields)  # unified: header row carries `::record_action`
    for i in range(max(len(v) for v in _UNI_ENUMS.values())):
        dl.append([
            _UNI_ENUMS[f][i] if i < len(_UNI_ENUMS[f]) else None
            for f in enum_fields
        ])
    wb.save(path)


@pytest.fixture
def unified_template(tmp_path):
    p = tmp_path / 'unified.xlsx'
    _make_unified_template(str(p))
    return str(p)


def _read_unified_rows(path):
    wb = openpyxl.load_workbook(path)
    ws = wb[listing_bulk.TEMPLATE_SHEET]
    names = [c.value for c in ws[5]]  # field API names at row 5
    idx = {n: i for i, n in enumerate(names)}
    rows = []
    for r in ws.iter_rows(min_row=6, values_only=True):
        if any(c is not None and str(c).strip() for c in r):
            rows.append({n: r[idx[n]] for n in names if n})
    return rows


def test_unified_header_and_dialect_detected(unified_template):
    wb = openpyxl.load_workbook(unified_template)
    ws = wb[listing_bulk.TEMPLATE_SHEET]
    # Field-name row is row 5 (unified), found structurally -- NOT item_sku.
    assert listing_bulk._find_header_row(ws) == 5
    cols = listing_bulk._field_columns(ws, 5)
    schema = listing_bulk._Schema(cols)
    assert schema.dialect == 'unified'
    # Friendly roles resolve to the decorated unified columns.
    assert schema.field('sku') == _UNI_SKU
    assert schema.field('operation') == _UNI_OP
    assert schema.field('parentage') == _UNI_PARENTAGE
    assert schema.field('parent_sku') == _UNI_PARENT_SKU
    assert schema.field('variation_theme') == _UNI_THEME
    assert schema.field('brand') == _UNI_BRAND
    assert schema.field('product_id') == _UNI_ID_VALUE
    assert schema.field('product_id_type') == _UNI_ID_TYPE
    # Required fields are parsed from the unified DD (6-col layout,
    # 'Required?' located by header name, not a fixed index).
    assert listing_bulk._load_required_fields(wb) == _UNI_REQUIRED


def test_fill_unified_clears_prefilled_rows_and_maps_friendly_keys(
    unified_template, tmp_path
):
    """The core regression: the SAME friendly spec that drives legacy must
    drive unified -- clearing Amazon's prefilled Example/instruction rows
    (the "1/8" cause) and routing every field to its decorated column."""
    spec = {
        'product_type': 'socks',
        'brand': 'ACME',
        'marketplace': 'SA',
        'rows': [
            {
                'sku': 'WIDGET-006',
                'operation': 'create',
                'parentage': 'Parent',
                'variation_theme': 'COLOR',
                'fields': {'item_name': 'ACME Socks'},
            },
            {
                'sku': 'WIDGET-006-WHT',
                'operation': 'create',
                'parentage': 'Child',
                'parent_sku': 'WIDGET-006',
                'variation_theme': 'COLOR',
                'fields': {
                    'item_name': 'ACME Socks White',
                    'our_price': '42.99',
                    'quantity': '100',
                },
            },
        ],
    }
    out = str(tmp_path / 'out.xlsx')
    _run([
        'fill',
        unified_template,
        '--spec',
        _spec(tmp_path, spec),
        '--out',
        out,
    ])
    rows = _read_unified_rows(out)
    skus = [r[_UNI_SKU] for r in rows]
    # Prefilled Example + "do not delete" rows are GONE; only the spec rows.
    assert 'EXAMPLE-SKU' not in skus
    assert not any('do not delete' in str(r[_UNI_SKU]) for r in rows)
    assert skus == ['WIDGET-006', 'WIDGET-006-WHT']
    parent = next(r for r in rows if r[_UNI_SKU] == 'WIDGET-006')
    child = next(r for r in rows if r[_UNI_SKU] == 'WIDGET-006-WHT')
    # create -> blank operation cell; friendly keys hit unified columns.
    assert parent[_UNI_OP] in (None, '')
    assert parent[_UNI_PARENTAGE] == 'Parent'
    assert parent[_UNI_PARENT_SKU] in (None, '')  # parent has no parent
    assert parent[_UNI_BRAND] == 'ACME'  # top-level brand default applied
    assert parent[_UNI_PRICE] in (None, '')  # offer is child-level
    assert child[_UNI_PARENTAGE] == 'Child'
    assert child[_UNI_PARENT_SKU] == 'WIDGET-006'
    assert child[_UNI_THEME] == 'COLOR'
    # bare our_price routed to the `[audience=ALL]` offer column...
    assert str(child[_UNI_PRICE]) == '42.99'
    # ...and bare quantity to the marketplace's fulfillment group.
    assert str(child[_UNI_QTY]) == '100'


def test_fill_unified_operation_tokens(unified_template, tmp_path):
    spec = {
        'product_type': 'socks',
        'brand': 'ACME',
        'marketplace': 'SA',
        'rows': [
            {'sku': 'W-1', 'operation': 'create', 'fields': {'item_name': 'x'}},
            {'sku': 'W-2', 'operation': 'update', 'fields': {'item_name': 'x'}},
            {
                'sku': 'W-3',
                'operation': 'partialupdate',
                'fields': {'item_name': 'x'},
            },
            {'sku': 'W-4', 'operation': 'delete'},
        ],
    }
    out = str(tmp_path / 'out.xlsx')
    _run([
        'fill',
        unified_template,
        '--spec',
        _spec(tmp_path, spec),
        '--out',
        out,
    ])
    by_sku = {r[_UNI_SKU]: r for r in _read_unified_rows(out)}
    assert by_sku['W-1'][_UNI_OP] in (None, '')  # create = blank
    assert by_sku['W-2'][_UNI_OP] == 'Create or Replace (Full Update)'
    assert by_sku['W-3'][_UNI_OP] == 'Edit (Partial Update)'
    assert by_sku['W-4'][_UNI_OP] == 'Delete'


def test_fill_unified_asin_folds_into_volt_product_id(
    unified_template, tmp_path
):
    spec = {
        'product_type': 'socks',
        'brand': 'ACME',
        'marketplace': 'SA',
        'rows': [
            {
                'sku': 'W-1',
                'operation': 'update',
                'asin': 'B0EXAMPLE1',
                'fields': {'item_name': 'x'},
            },
        ],
    }
    out = str(tmp_path / 'out.xlsx')
    _run([
        'fill',
        unified_template,
        '--spec',
        _spec(tmp_path, spec),
        '--out',
        out,
    ])
    row = _read_unified_rows(out)[0]
    assert row[_UNI_ID_VALUE] == 'B0EXAMPLE1'
    # Canonicalised to the template's exact enum case (unified lists ASIN,
    # not asin) -- Amazon is case-strict on some fields.
    assert row[_UNI_ID_TYPE] == 'ASIN'


def test_fill_unified_tsv_keeps_settings_header(unified_template, tmp_path):
    """The uploaded .txt must retain the row-1 settings/signature block --
    Amazon's introspect-feed keys on it to detect the file type. A
    hand-rolled header-less TSV was rejected ("File upload unsuccessful")."""
    spec = {
        'product_type': 'socks',
        'brand': 'ACME',
        'marketplace': 'SA',
        'rows': [
            {'sku': 'W-1', 'operation': 'create', 'fields': {'item_name': 'x'}}
        ],
    }
    out = str(tmp_path / 'out.xlsx')
    _run([
        'fill',
        unified_template,
        '--spec',
        _spec(tmp_path, spec),
        '--out',
        out,
    ])
    txt = tmp_path / 'out.txt'
    lines = [ln for ln in txt.read_text(encoding='utf-8').split('\n') if ln]
    grid = [ln.split('\t') for ln in lines]
    assert grid[0][0].startswith('settings=')  # detection header preserved
    assert _UNI_SKU in grid[4]  # field-name row at index 4 (Excel row 5)
    assert len({len(r) for r in grid}) == 1  # uniform width, no drift
    # Prefilled Example row did not survive into the upload artefact.
    assert not any('EXAMPLE-SKU' in r for r in grid)


def test_fill_unified_writes_data_at_datarow(unified_template, tmp_path):
    """Unified data must start at the settings' `dataRow` (8 here), NOT
    header_row+1. Amazon SKIPS the rows between the field-name row and
    dataRow as an example region, so data written there is dropped
    (verified live: children written into rows 6-7 processed as 4/6)."""
    spec = {
        'product_type': 'socks',
        'brand': 'ACME',
        'marketplace': 'SA',
        'rows': [
            {
                'sku': 'P',
                'operation': 'create',
                'parentage': 'Parent',
                'variation_theme': 'COLOR',
                'fields': {'item_name': 'p'},
            },
            {
                'sku': 'C',
                'operation': 'create',
                'parentage': 'Child',
                'parent_sku': 'P',
                'variation_theme': 'COLOR',
                'fields': {
                    'item_name': 'c',
                    'our_price': '9.99',
                    'quantity': '5',
                },
            },
        ],
    }
    out = str(tmp_path / 'out.xlsx')
    _run([
        'fill',
        unified_template,
        '--spec',
        _spec(tmp_path, spec),
        '--out',
        out,
    ])
    ws = openpyxl.load_workbook(out)[listing_bulk.TEMPLATE_SHEET]
    # The example region (rows 6-7) is blank; data starts at row 8.
    assert all(c.value in (None, '') for c in ws[6])
    assert all(c.value in (None, '') for c in ws[7])
    assert ws.cell(row=8, column=1).value == 'P'  # first data row = 8
    assert ws.cell(row=9, column=1).value == 'C'


def test_parse_feedback_unified_report_finds_child_error(tmp_path, capsys):
    """The blind spot that caused the 40-minute misdiagnosis: a unified
    processing report must be parseable so the per-CHILD rejection reason
    is surfaced (it was invisible because parse-feedback also keyed on
    item_sku)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = listing_bulk.TEMPLATE_SHEET
    ws.append(['settings=feedType=256'])  # row 1
    ws.append(['instructions'])  # row 2
    ws.append(['group'])  # row 3
    ws.append(['SKU', 'Colour'])  # row 4 localised labels
    ws.append([_UNI_SKU, _UNI_COLOR])  # row 5 field API names
    ws.append(['WIDGET-006-WHT', 'White'])  # row 6 data
    ws.cell(row=6, column=1).comment = Comment(
        'ERROR : 8560 does not match any ASINs; include a valid '
        'standard_product_id or request GTIN exemption',
        'Amazon',
    )
    p = tmp_path / 'report.xlsx'
    wb.save(str(p))

    errs = list(listing_bulk._template_cell_errors(str(p)))
    assert errs and all(sku == 'WIDGET-006-WHT' for sku, _, _ in errs)
    assert any('8560' in m for _, _, m in errs)

    with pytest.raises(SystemExit):  # exit 1 -- a blocking error exists
        _run(['parse-feedback', str(p)])
    out = capsys.readouterr().out
    assert 'WIDGET-006-WHT' in out and '8560' in out


class TestReviewContractTargetMarketplace:
    """The DoD review contract must require verifying the listing on the
    TARGET marketplace — the design fix for the SA/AE confusion where a
    listing "for AE" completed on an account-level batch that actually
    showed on SA. Must cover BOTH account structures (unified pan-regional
    and separate per-marketplace)."""

    def _review(self):
        skill_md = (
            Path(__file__).resolve().parents[2]
            / 'app'
            / 'skills_v2'
            / 'amazon-listing'
            / 'SKILL.md'
        )
        review = parse_skill_review(skill_md)
        assert review is not None, 'amazon-listing must declare a review block'
        return (review.criteria + '\n' + review.verify_by).lower()

    def test_requires_target_marketplace_verification(self):
        text = self._review()
        assert 'target' in text and 'marketplace' in text
        # Verify on the target marketplace's OWN inventory, not a batch row.
        assert 'manage inventory' in text or 'inventory' in text
        assert 'batch' in text  # the account-level batch trap is named

    def test_covers_both_account_structures(self):
        text = self._review()
        assert 'unified' in text, 'must name the unified pan-regional case'
        assert 'separate' in text, 'must name the separate per-marketplace case'

    def test_wrong_marketplace_is_a_gap(self):
        text = self._review()
        # A listing on the wrong marketplace / only a batch id must not pass.
        assert 'gap' in text
        assert 'wrong marketplace' in text or 'different marketplace' in text


# --- region-stamp guard (wrong-marketplace template must not fill) ---


def test_fill_hard_fails_on_wrong_region_template(mkt_template, tmp_path):
    # The live incident: the store tick silently failed, the template
    # came back stamped for a different marketplace, and the upload
    # then "succeeded" on the wrong storefront. Declaring the target
    # must make that a hard error at fill time.
    spec = _spec(
        tmp_path,
        {
            'marketplace': 'EG',  # not stamped in this SA+AE template
            'product_type': 'socks',
            'rows': [
                {
                    'sku': 'K-WHT',
                    'parentage': 'Child',
                    'fields': {
                        'item_name': 'x',
                        'feed_product_type': 'socks',
                    },
                }
            ],
        },
    )
    out = str(tmp_path / 'out.xlsx')
    with pytest.raises(SystemExit) as e:
        _run(['fill', mkt_template, '--spec', spec, '--out', out])
    msg = str(e.value)
    assert 'region-stamped' in msg and 'regenerate' in msg


def test_fill_cli_marketplace_flag_also_guarded(mkt_template, tmp_path):
    spec = _spec(
        tmp_path,
        {
            'product_type': 'socks',
            'rows': [
                {
                    'sku': 'K-WHT',
                    'parentage': 'Child',
                    'fields': {
                        'item_name': 'x',
                        'feed_product_type': 'socks',
                    },
                }
            ],
        },
    )
    out = str(tmp_path / 'out.xlsx')
    with pytest.raises(SystemExit) as e:
        _run([
            'fill',
            mkt_template,
            '--spec',
            spec,
            '--out',
            out,
            '--marketplace',
            'EG',
        ])
    assert 'region-stamped' in str(e.value)


def test_fill_auto_adopt_single_stamp_warns_loudly(template, tmp_path, capsys):
    # No declared marketplace + single-stamped template: auto-adopt is
    # kept (legacy flows), but the adopted storefront must be shouted.
    spec = _spec(
        tmp_path,
        {
            'product_type': 'socks',
            'brand': 'acme',
            'rows': [
                {
                    'sku': 'K-WHT',
                    'fields': {
                        'item_name': 'x',
                        'feed_product_type': 'socks',
                    },
                }
            ],
        },
    )
    out = str(tmp_path / 'out.xlsx')
    _run(['fill', template, '--spec', spec, '--out', out])
    err = capsys.readouterr().err
    assert 'auto-adopting' in err and 'MKTSA' in err


def test_inspect_prints_region_stamp(mkt_template, capsys):
    _run(['inspect', mkt_template])
    out = capsys.readouterr().out
    assert 'marketplaces:' in out
    assert 'A17E79C6D8DWNP (SA)' in out and 'A2VIGQ35RCS4UG (AE)' in out
